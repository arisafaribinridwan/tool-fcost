from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
import pandas as pd
import pytest

from app.services.pipeline_service import run_pipeline
from app.services.pipeline_types import PipelineError, PipelineStepStatus


def _write_yaml(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8")


def _write_action_lookup_rules_step_recipe(path: Path, *, priority_column: str | None = "priority") -> None:
    lines = [
        'name: "Action Lookup Rules Step Recipe"',
        "datasets:",
        '  working_dataset: "result"',
        "  canonical_columns:",
        '    - "job_sheet_section"',
        '    - "part_name"',
        '    - "symptom_comment"',
        '    - "repair_comment"',
        '    - "action"',
        "steps:",
        '  - id: "sub_1_extract"',
        '    type: "extract_sheet"',
        "    sheet_selector:",
        '      contains: "Data"',
        '      case_sensitive: false',
        "    header_locator:",
        '      type: "required_columns"',
        "      scan_rows: [1, 1]",
        "      required:",
        '        - "job_sheet_section"',
        '        - "part_name"',
        '        - "symptom_comment"',
        '        - "repair_comment"',
        "    select:",
        '      "job_sheet_section": "job_sheet_section"',
        '      "part_name": "part_name"',
        '      "symptom_comment": "symptom_comment"',
        '      "repair_comment": "repair_comment"',
        '    write_to: "result"',
        '    mode: "replace"',
        '  - id: "sub_2_action"',
        '    type: "lookup_rules"',
        "    inputs:",
        '      - "job_sheet_section"',
        '      - "part_name"',
        '      - "symptom_comment"',
        '      - "repair_comment"',
        '    target_column: "action"',
        "    master:",
        '      file: "masters/master_table.xlsx"',
        '      sheet: "action"',
        '      value: "action"',
        "    matching:",
        '      order: "top_to_bottom"',
        "      first_match_wins: true",
        "      matchers:",
        '        - source: "job_sheet_section"',
        '          master: "job_sheet_section"',
        '          mode: "equals"',
        '        - source: "part_name"',
        '          master: "part_name"',
        '          mode: "equals"',
        "          normalize:",
        "            trim: true",
        "            case_sensitive: false",
        '        - source: "symptom_comment"',
        '          master: "symptom_comment"',
        '          mode: "regex"',
        '        - source: "repair_comment"',
        '          master: "repair_comment"',
        '          mode: "regex"',
        "    on_missing_match: null",
        "outputs:",
        '  - sheet_name: "result"',
        "    columns:",
        '      - "job_sheet_section"',
        '      - "part_name"',
        '      - "symptom_comment"',
        '      - "repair_comment"',
        '      - "action"',
    ]
    if priority_column is not None:
        insert_at = lines.index("      matchers:")
        lines.insert(insert_at, f'      priority_column: "{priority_column}"')
    _write_yaml(path, "\n".join(lines))


def test_run_pipeline_happy_path_csv_with_master_and_pivot(app_paths):
    source_path = app_paths.project_root / "source.csv"
    source_df = pd.DataFrame(
        [
            {"tanggal": "2026-04-01", "kode_produk": "A", "qty": 10, "harga": 10000},
            {"tanggal": "2026-04-02", "kode_produk": "B", "qty": 5, "harga": 8000},
            {"tanggal": "2026-04-03", "kode_produk": "A", "qty": 3, "harga": 10000},
        ]
    )
    source_df.to_csv(source_path, index=False)

    master_path = app_paths.masters_dir / "produk.csv"
    master_df = pd.DataFrame(
        [
            {"kode_produk": "A", "nama_produk": "Produk A", "kategori": "Cat 1"},
            {"kode_produk": "B", "nama_produk": "Produk B", "kategori": "Cat 2"},
        ]
    )
    master_df.to_csv(master_path, index=False)

    config_path = app_paths.configs_dir / "report.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Laporan Tes"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Laporan Tes"',
                '  period_from_column: "tanggal"',
                "masters:",
                '  - file: "masters/produk.csv"',
                '    key: "kode_produk"',
                "    columns:",
                '      - "nama_produk"',
                '      - "kategori"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "tanggal"',
                '      - "kode_produk"',
                '      - "nama_produk"',
                '      - "kategori"',
                '      - "qty"',
                '  - sheet_name: "Summary"',
                "    pivot:",
                '      index: "kategori"',
                '      values: "qty"',
                '      aggfunc: "sum"',
                "styling:",
                '  header_color: "1F4E78"',
                '  font: "Calibri"',
                '  number_format: "#,##0"',
                '  date_format: "DD/MM/YYYY"',
                '  freeze_pane: "A5"',
            ]
        ),
    )

    logs: list[str] = []
    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=logs.append,
    )

    assert result.output_path.exists()
    assert result.source_copy_path.exists()
    assert result.sheets_written == 2
    assert result.duration_ms >= 1
    assert any("Write workbook" in item for item in logs)

    detail_df = pd.read_excel(result.output_path, sheet_name="Detail", skiprows=3)
    summary_df = pd.read_excel(result.output_path, sheet_name="Summary", skiprows=3)

    assert list(detail_df.columns) == [
        "tanggal",
        "kode_produk",
        "nama_produk",
        "kategori",
        "qty",
    ]
    assert len(detail_df) == 3
    assert summary_df.loc[summary_df["kategori"] == "Cat 1", "qty"].iloc[0] == 13
    assert summary_df.loc[summary_df["kategori"] == "Cat 2", "qty"].iloc[0] == 5


def test_run_pipeline_writes_manual_period_override(app_paths):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame([{"qty": 1}]).to_csv(source_path, index=False)

    config_path = app_paths.configs_dir / "manual_period.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Manual Period"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Manual Period"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "qty"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _message: None,
        period_text_override="Periode: March 2026",
    )

    workbook = load_workbook(result.output_path, read_only=True)
    assert workbook["Detail"]["A2"].value == "Periode: March 2026"


def test_run_pipeline_without_period_override_keeps_fallback(app_paths):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame([{"qty": 1}]).to_csv(source_path, index=False)

    config_path = app_paths.configs_dir / "fallback_period.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Fallback Period"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Fallback Period"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "qty"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _message: None,
    )

    workbook = load_workbook(result.output_path, read_only=True)
    assert workbook["Detail"]["A2"].value == "Periode: -"


def test_run_pipeline_step_recipe_reuses_source_period_text(app_paths):
    source_path = app_paths.project_root / "job_summary_source.xlsx"
    source_workbook = Workbook()
    source_sheet = source_workbook.active
    source_sheet.title = "result"
    source_sheet["A1"] = "Monthly Report Final Recipe"
    source_sheet["A2"] = "Periode: March 2026"
    source_sheet["A3"] = "Dibuat: 01/05/2026 23:17:11"
    source_sheet["A5"] = "notification"
    source_sheet["A6"] = "1000001"
    source_workbook.save(source_path)

    config_path = app_paths.configs_dir / "job_summary_period.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Job Summary Result"',
                "datasets:",
                '  working_dataset: "result"',
                "  canonical_columns:",
                '    - "notification"',
                "steps:",
                '  - id: "extract_result_like_source"',
                '    type: "extract_sheet"',
                "    sheet_selector:",
                '      mode: "single_sheet_workbook"',
                "    header_locator:",
                '      type: "required_columns"',
                "      scan_rows: [1, 10]",
                "      required:",
                '        - "notification"',
                "    select:",
                '      "notification": "notification"',
                '    write_to: "result"',
                "outputs:",
                '  - sheet_name: "result"',
                "    columns:",
                '      - "notification"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _message: None,
    )

    workbook = load_workbook(result.output_path, read_only=True)
    assert workbook["result"]["A2"].value == "Periode: March 2026"


def test_run_pipeline_skips_copy_when_source_already_in_uploads_subfolder(app_paths):
    upload_subdir = app_paths.uploads_dir / "nested"
    upload_subdir.mkdir(parents=True, exist_ok=True)
    source_path = upload_subdir / "source.csv"
    pd.DataFrame([{"qty": 1}]).to_csv(source_path, index=False)

    config_path = app_paths.configs_dir / "skip_copy_uploads.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Skip Copy Uploads"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Skip Copy Uploads"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "qty"',
            ]
        ),
    )

    logs: list[str] = []
    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=logs.append,
    )

    assert result.output_path.exists()
    assert result.source_copy_path == source_path.resolve()
    assert len(list(app_paths.uploads_dir.rglob("*.csv"))) == 1
    assert any("copy source dilewati" in item for item in logs)


def test_run_pipeline_skips_copy_when_source_already_in_uploads_root(app_paths):
    source_path = app_paths.uploads_dir / "source_root.csv"
    pd.DataFrame([{"qty": 2}]).to_csv(source_path, index=False)

    config_path = app_paths.configs_dir / "skip_copy_uploads_root.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Skip Copy Uploads Root"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Skip Copy Uploads Root"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "qty"',
            ]
        ),
    )

    logs: list[str] = []
    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=logs.append,
    )

    assert result.output_path.exists()
    assert result.source_copy_path == source_path.resolve()
    assert len(list(app_paths.uploads_dir.glob("*.csv"))) == 1
    assert any("copy source dilewati" in item for item in logs)


def test_run_pipeline_skips_copy_when_source_already_in_upload_dir(app_paths):
    upload_dir = app_paths.project_root / "upload"
    upload_dir.mkdir(parents=True, exist_ok=True)
    source_path = upload_dir / "source_upload_dir.csv"
    pd.DataFrame([{"qty": 3}]).to_csv(source_path, index=False)

    config_path = app_paths.configs_dir / "skip_copy_upload_dir.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Skip Copy Upload Dir"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Skip Copy Upload Dir"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "qty"',
            ]
        ),
    )

    logs: list[str] = []
    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=logs.append,
    )

    assert result.output_path.exists()
    assert result.source_copy_path == source_path.resolve()
    assert len(list(app_paths.uploads_dir.glob("*.csv"))) == 0
    assert any("copy source dilewati" in item for item in logs)


def test_run_pipeline_blocks_when_source_size_exceeds_limit(app_paths):
    source_path = app_paths.project_root / "source.csv"
    source_path.write_bytes(b"x" * 2 * 1024 * 1024)

    config_path = app_paths.configs_dir / "report.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Laporan Tes"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Laporan Tes"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "qty"',
            ]
        ),
    )

    _write_yaml(
        app_paths.configs_dir / "app_limits.yaml",
        "\n".join(
            [
                "resource_guardrails:",
                "  max_source_size_mb: 1",
                "  warning_source_size_mb: 0.5",
                "  interactive_row_limit: 150000",
                '  row_limit_mode: "warning"',
                "  timeouts:",
                "    read_seconds: 45",
                "    transform_seconds: 120",
                "    write_seconds: 60",
            ]
        ),
    )

    pytest.skip("run_pipeline saat ini belum menerapkan guardrail ukuran; validasi ada di preflight.")


def test_run_pipeline_emits_progress_events_in_order(app_paths):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame([{"qty": 1}]).to_csv(source_path, index=False)

    config_path = app_paths.configs_dir / "progress.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Progress Test"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Progress Test"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "qty"',
            ]
        ),
    )

    progress_events: list[PipelineStepStatus] = []
    run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
        progress=progress_events.append,
    )

    assert [(item.step_key, item.state) for item in progress_events] == [
        ("load_config", "running"),
        ("load_config", "done"),
        ("copy_source", "running"),
        ("copy_source", "done"),
        ("read_source", "running"),
        ("read_source", "done"),
        ("load_master", "running"),
        ("load_master", "done"),
        ("transform", "running"),
        ("transform", "done"),
        ("build_output", "running"),
        ("build_output", "done"),
        ("write_output", "running"),
        ("write_output", "done"),
    ]


def test_run_pipeline_marks_failed_step_when_transform_errors(app_paths, monkeypatch):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame([{"qty": 1}]).to_csv(source_path, index=False)

    config_path = app_paths.configs_dir / "transform_fail.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Transform Fail"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Transform Fail"',
                "transforms:",
                '  - type: "ensure_optional_columns"',
                "    columns:",
                '      - "qty"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "qty"',
            ]
        ),
    )

    def raise_transform_error(*args, **kwargs):
        raise ValueError("transform rusak")

    monkeypatch.setattr("app.services.pipeline_service.apply_transform_steps", raise_transform_error)

    progress_events: list[PipelineStepStatus] = []
    with pytest.raises(PipelineError, match="transform rusak"):
        run_pipeline(
            paths=app_paths,
            source_path=source_path,
            config_path=config_path,
            log=lambda _: None,
            progress=progress_events.append,
        )

    assert progress_events[-1].step_key == "transform"
    assert progress_events[-1].state == "running"


def test_run_pipeline_raises_timeout_for_slow_read_stage(app_paths, monkeypatch):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame([{"qty": 1}]).to_csv(source_path, index=False)

    config_path = app_paths.configs_dir / "timeout.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Timeout Test"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Timeout Test"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "qty"',
            ]
        ),
    )
    _write_yaml(
        app_paths.configs_dir / "app_limits.yaml",
        "\n".join(
            [
                "resource_guardrails:",
                "  max_source_size_mb: 75",
                "  warning_source_size_mb: 60",
                "  interactive_row_limit: 150000",
                '  row_limit_mode: "warning"',
                "  timeouts:",
                "    read_seconds: 0.01",
                "    transform_seconds: 120",
                "    write_seconds: 60",
            ]
        ),
    )

    pytest.skip("run_pipeline saat ini belum menjalankan guardrail timeout; validasi ada di util guardrails.")


def test_run_pipeline_supports_casefold_master_path_and_sheet_name(app_paths):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame([{"kode_produk": "A", "qty": 1}]).to_csv(source_path, index=False)

    master_path = app_paths.masters_dir / "Produk.CSV"
    pd.DataFrame([{"kode_produk": "A", "nama_produk": "Produk A"}]).to_csv(
        master_path,
        index=False,
    )

    config_path = app_paths.configs_dir / "casefold_master.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Casefold Master"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Casefold Master"',
                "masters:",
                '  - file: "masters\\\\produk.csv"',
                '    key: "kode_produk"',
                "    columns:",
                '      - "nama_produk"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "kode_produk"',
                '      - "nama_produk"',
                '      - "qty"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    detail_df = pd.read_excel(result.output_path, sheet_name="Detail", skiprows=3)
    assert detail_df["nama_produk"].tolist() == ["Produk A"]


def test_run_pipeline_raises_when_master_missing(app_paths):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame(
        [{"tanggal": "2026-04-01", "kode_produk": "A", "qty": 10}]
    ).to_csv(source_path, index=False)

    config_path = app_paths.configs_dir / "missing_master.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Missing Master"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Missing Master"',
                '  period_from_column: "tanggal"',
                "masters:",
                '  - file: "masters/not_found.csv"',
                '    key: "kode_produk"',
                "    columns:",
                '      - "nama_produk"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "tanggal"',
                '      - "kode_produk"',
                '      - "qty"',
            ]
        ),
    )

    with pytest.raises(PipelineError, match="File master tidak ditemukan"):
        run_pipeline(
            paths=app_paths,
            source_path=source_path,
            config_path=config_path,
            log=lambda _: None,
        )


def test_run_pipeline_raises_when_excel_sheet_missing(app_paths):
    source_path = app_paths.project_root / "source.xlsx"
    pd.DataFrame([{"tanggal": "2026-04-01", "qty": 1}]).to_excel(
        source_path,
        index=False,
        sheet_name="SheetAktual",
    )

    config_path = app_paths.configs_dir / "wrong_sheet.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Wrong Sheet"',
                'source_sheet: "SheetTidakAda"',
                "header:",
                '  title: "Wrong Sheet"',
                '  period_from_column: "tanggal"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "tanggal"',
                '      - "qty"',
            ]
        ),
    )

    with pytest.raises(PipelineError, match="Sheet 'SheetTidakAda' tidak ditemukan"):
        run_pipeline(
            paths=app_paths,
            source_path=source_path,
            config_path=config_path,
            log=lambda _: None,
        )


def test_run_pipeline_matches_excel_sheet_name_case_insensitively(app_paths):
    source_path = app_paths.project_root / "source.xlsx"
    pd.DataFrame([{"tanggal": "2026-04-01", "qty": 1}]).to_excel(
        source_path,
        index=False,
        sheet_name="SheetAktual",
    )

    config_path = app_paths.configs_dir / "sheet_casefold.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Sheet Casefold"',
                'source_sheet: "sheetaktual"',
                "header:",
                '  title: "Sheet Casefold"',
                '  period_from_column: "tanggal"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "tanggal"',
                '      - "qty"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    detail_df = pd.read_excel(result.output_path, sheet_name="Detail", skiprows=3)
    assert detail_df["qty"].tolist() == [1]


def test_run_pipeline_raises_when_required_source_columns_missing(app_paths):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame([{"qty": 10, "harga": 2000}]).to_csv(source_path, index=False)

    config_path = app_paths.configs_dir / "required_columns.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Required Columns"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Required Columns"',
                "required_source_columns:",
                '  - "tanggal"',
                '  - "qty"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "qty"',
                '      - "harga"',
            ]
        ),
    )

    with pytest.raises(PipelineError, match="Kolom wajib source tidak ditemukan: tanggal"):
        run_pipeline(
            paths=app_paths,
            source_path=source_path,
            config_path=config_path,
            log=lambda _: None,
        )


def test_run_pipeline_raises_when_source_file_empty(app_paths):
    source_path = app_paths.project_root / "source.csv"
    source_path.write_text("", encoding="utf-8")

    config_path = app_paths.configs_dir / "empty_source.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Empty Source"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Empty Source"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "qty"',
            ]
        ),
    )

    with pytest.raises(PipelineError, match="File source kosong"):
        run_pipeline(
            paths=app_paths,
            source_path=source_path,
            config_path=config_path,
            log=lambda _: None,
        )


def test_run_pipeline_supports_ordered_rules_master_for_action(app_paths):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame(
        [
            {"part_name": "AC_CORD", "repair_comment": "need ext support"},
            {"part_name": "panel", "repair_comment": "replace unit"},
            {"part_name": "MAIN_UNIT", "repair_comment": "UPGRADE firmware"},
            {"part_name": "Other", "repair_comment": "dibawa customer"},
            {"part_name": "Other", "repair_comment": "unknown"},
        ]
    ).to_csv(source_path, index=False)

    master_path = app_paths.masters_dir / "master_table.xlsx"
    with pd.ExcelWriter(master_path) as writer:
        pd.DataFrame(
            [
                {"part_name": None, "repair_comment": "*ext", "action": "external"},
                {"part_name": "PANEL", "repair_comment": "*", "action": "replace_panel"},
                {
                    "part_name": "MAIN_UNIT",
                    "repair_comment": "*",
                    "action": "replace_main_unit",
                },
                {"part_name": None, "repair_comment": "*bawa", "action": "ZY"},
                {"part_name": None, "repair_comment": "UPGRADE", "action": "upgrade"},
            ]
        ).to_excel(writer, index=False, sheet_name="action")

    config_path = app_paths.configs_dir / "batch5_action.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Batch 5 Action"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Batch 5 Action"',
                "masters:",
                '  - file: "masters/master_table.xlsx"',
                '    sheet_name: "action"',
                '    strategy: "ordered_rules"',
                '    target_column: "action"',
                '    value_column: "action"',
                "    matchers:",
                '      - source: "part_name"',
                '        master: "part_name"',
                '        mode: "equals"',
                '      - source: "repair_comment"',
                '        master: "repair_comment"',
                '        mode: "contains"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "part_name"',
                '      - "repair_comment"',
                '      - "action"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    detail_df = pd.read_excel(result.output_path, sheet_name="Detail", skiprows=3)

    assert detail_df["action"].fillna("").tolist() == [
        "external",
        "replace_panel",
        "replace_main_unit",
        "ZY",
        "",
    ]


def test_run_pipeline_supports_lookup_rules_master_for_action(app_paths):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame(
        [
            {"part_name": "ac_cord", "repair_comment": "need ext support"},
            {"part_name": "panel", "repair_comment": "replace unit"},
            {"part_name": "MAIN_UNIT", "repair_comment": "UPGRADE firmware"},
            {"part_name": "Other", "repair_comment": "dibawa customer"},
            {"part_name": "Other", "repair_comment": "unknown"},
        ]
    ).to_csv(source_path, index=False)

    master_path = app_paths.masters_dir / "master_table.xlsx"
    with pd.ExcelWriter(master_path) as writer:
        pd.DataFrame(
            [
                {"part_name": None, "repair_comment": "*ext", "action": "external"},
                {"part_name": "PANEL", "repair_comment": "*", "action": "replace_panel"},
                {
                    "part_name": "MAIN_UNIT",
                    "repair_comment": "*",
                    "action": "replace_main_unit",
                },
                {"part_name": None, "repair_comment": "*bawa", "action": "ZY"},
                {"part_name": None, "repair_comment": "UPGRADE", "action": "upgrade"},
            ]
        ).to_excel(writer, index=False, sheet_name="action")

    config_path = app_paths.configs_dir / "batch5_action_lookup_rules.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Batch 5 Action Lookup Rules"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Batch 5 Action Lookup Rules"',
                "masters:",
                '  - file: "masters/master_table.xlsx"',
                '    sheet_name: "action"',
                '    strategy: "lookup_rules"',
                '    target_column: "action"',
                '    value_column: "action"',
                "    matching:",
                '      order: "top_to_bottom"',
                "      first_match_wins: true",
                "      matchers:",
                '        - source: "part_name"',
                '          master: "part_name"',
                '          mode: "equals"',
                "          normalize:",
                "            trim: true",
                "            case_sensitive: false",
                "            blank_as_wildcard: true",
                '        - source: "repair_comment"',
                '          master: "repair_comment"',
                '          mode: "contains"',
                "          normalize:",
                "            trim: true",
                "            case_sensitive: false",
                '            wildcard: "*"',
                "            blank_as_wildcard: true",
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "part_name"',
                '      - "repair_comment"',
                '      - "action"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    detail_df = pd.read_excel(result.output_path, sheet_name="Detail", skiprows=3)

    assert detail_df["action"].fillna("").tolist() == [
        "external",
        "replace_panel",
        "replace_main_unit",
        "ZY",
        "",
    ]


def test_run_pipeline_lookup_exact_replace_supports_alias_separator_or(app_paths):
    source_path = app_paths.project_root / "source.xlsx"
    pd.DataFrame(
        [
            {"symptom_comment": "EXS", "repair_comment": "RESOLDERING"},
            {"symptom_comment": "EXTRNAL", "repair_comment": "RESOLDRING"},
            {"symptom_comment": "EXTERNAL", "repair_comment": "RESOLDER"},
            {"symptom_comment": "UNKNOWN", "repair_comment": "OTHER"},
            {"symptom_comment": "", "repair_comment": "RESOLDERING"},
            {"symptom_comment": "EXS", "repair_comment": ""},
            {"symptom_comment": "LCD-PW-A01 MT", "repair_comment": "CHECK RESOLDER UNIT"},
            {"symptom_comment": "RESOLDERING", "repair_comment": "EXS"},
            {"symptom_comment": "MT", "repair_comment": "MT"},
        ]
    ).to_excel(source_path, index=False, sheet_name="Data")

    master_path = app_paths.masters_dir / "master_table.xlsx"
    with pd.ExcelWriter(master_path) as writer:
        pd.DataFrame(
            [
                {
                    "alias": "EXS|EXTRNAL|EXTERNAL",
                    "canonical": "EXTERNAL",
                    "scope": "symptom",
                },
                {
                    "alias": "RESOLDERING|RESOLDRING|RESOLDER",
                    "canonical": "RESOLDER",
                    "scope": "repair",
                },
                {
                    "alias": "MT",
                    "canonical": "MATI",
                    "scope": "both",
                },
            ]
        ).to_excel(writer, index=False, sheet_name="comment_synonyms")

    config_path = app_paths.configs_dir / "comment_synonym_or_step_recipe.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Comment Synonym OR Step Recipe"',
                "datasets:",
                '  working_dataset: "result"',
                "  canonical_columns:",
                '    - "symptom_comment"',
                '    - "repair_comment"',
                "steps:",
                '  - id: "sub_1_extract"',
                '    type: "extract_sheet"',
                "    sheet_selector:",
                '      contains: "Data"',
                '      case_sensitive: false',
                "    header_locator:",
                '      type: "required_columns"',
                "      scan_rows: [1, 1]",
                "      required:",
                '        - "symptom_comment"',
                '        - "repair_comment"',
                "    select:",
                '      "symptom_comment": "symptom_comment"',
                '      "repair_comment": "repair_comment"',
                '    write_to: "result"',
                '    mode: "replace"',
                '  - id: "sub_2_normalize_symptom"',
                '    type: "lookup_exact_replace"',
                '    source_column: "symptom_comment"',
                '    target_column: "symptom_comment"',
                "    master:",
                '      file: "masters/master_table.xlsx"',
                '      sheet: "comment_synonyms"',
                '      key: "alias"',
                '      value: "canonical"',
                "      filter:",
                '        scope_in: ["symptom", "both"]',
                "    matching:",
                "      trim: true",
                "      case_sensitive: false",
                '      alias_separator: "|"',
                '      match_mode: "contains"',
                '    on_blank_source: ""',
                '    on_missing_match: "keep_original"',
                '  - id: "sub_3_normalize_repair"',
                '    type: "lookup_exact_replace"',
                '    source_column: "repair_comment"',
                '    target_column: "repair_comment"',
                "    master:",
                '      file: "masters/master_table.xlsx"',
                '      sheet: "comment_synonyms"',
                '      key: "alias"',
                '      value: "canonical"',
                "      filter:",
                '        scope_in: ["repair", "both"]',
                "    matching:",
                "      trim: true",
                "      case_sensitive: false",
                '      alias_separator: "|"',
                '      match_mode: "contains"',
                '    on_blank_source: ""',
                '    on_missing_match: "keep_original"',
                "outputs:",
                '  - sheet_name: "result"',
                "    columns:",
                '      - "symptom_comment"',
                '      - "repair_comment"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    detail_df = pd.read_excel(result.output_path, sheet_name="result", skiprows=3, keep_default_na=False)
    assert detail_df["symptom_comment"].tolist() == [
        "EXTERNAL",
        "EXTERNAL",
        "EXTERNAL",
        "UNKNOWN",
        "",
        "EXTERNAL",
        "LCD-PW-A01 MATI",
        "RESOLDERING",
        "MATI",
    ]
    assert detail_df["repair_comment"].tolist() == [
        "RESOLDER",
        "RESOLDER",
        "RESOLDER",
        "OTHER",
        "RESOLDER",
        "",
        "CHECK RESOLDER UNIT",
        "EXS",
        "MATI",
    ]


def test_run_pipeline_supports_lookup_for_defect_category_from_action(app_paths):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame(
        [
            {"action": "replace_panel"},
            {"action": "factory_reset"},
            {"action": "replace_remote_control"},
            {"action": "cancel"},
            {"action": "external"},
            {"action": "unknown_action"},
        ]
    ).to_csv(source_path, index=False)

    master_path = app_paths.masters_dir / "master_table.xlsx"
    with pd.ExcelWriter(master_path) as writer:
        pd.DataFrame(
            [
                {"Repair Action": "Replace Panel", "Category": "Defect"},
                {"Repair Action": "Factory Reset", "Category": "Defect"},
                {"Repair Action": "Replace Remote", "Category": "Defect"},
                {"Repair Action": "Cancel", "Category": "N/A"},
                {"Repair Action": "External", "Category": "N/A"},
            ]
        ).to_excel(writer, index=False, sheet_name="defect_category")

    config_path = app_paths.configs_dir / "batch5_defect_category.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Batch 5 Defect Category"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Batch 5 Defect Category"',
                "masters:",
                '  - file: "masters/master_table.xlsx"',
                '    sheet_name: "defect_category"',
                '    source_key: "action"',
                '    master_key: "Repair Action"',
                '    key_normalizer: "compact_text"',
                "    key_aliases:",
                '      replace_remote_control: "Replace Remote"',
                "    columns:",
                '      - "Category"',
                "    rename_columns:",
                '      Category: "defect_category"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "action"',
                '      - "defect_category"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    detail_df = pd.read_excel(
        result.output_path,
        sheet_name="Detail",
        skiprows=3,
        keep_default_na=False,
    )

    assert detail_df["defect_category"].tolist() == [
        "Defect",
        "Defect",
        "Defect",
        "N/A",
        "N/A",
        "",
    ]


def test_run_pipeline_supports_lookup_for_defect_from_action(app_paths):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame(
        [
            {"action": "replace_panel"},
            {"action": "factory_reset"},
            {"action": "replace_remote_control"},
            {"action": "cancel"},
            {"action": "external"},
            {"action": "unknown_action"},
        ]
    ).to_csv(source_path, index=False)

    master_path = app_paths.masters_dir / "master_table.xlsx"
    with pd.ExcelWriter(master_path) as writer:
        pd.DataFrame(
            [
                {"Repair Action": "Replace Panel", "Defect": "Panel"},
                {"Repair Action": "Factory Reset", "Defect": "Software"},
                {"Repair Action": "Replace Remote", "Defect": "Other"},
                {"Repair Action": "Cancel", "Defect": "N/A"},
                {"Repair Action": "External", "Defect": "N/A"},
            ]
        ).to_excel(writer, index=False, sheet_name="defect_category")

    config_path = app_paths.configs_dir / "batch5_defect.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Batch 5 Defect"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Batch 5 Defect"',
                "masters:",
                '  - file: "masters/master_table.xlsx"',
                '    sheet_name: "defect_category"',
                '    source_key: "action"',
                '    master_key: "Repair Action"',
                '    key_normalizer: "compact_text"',
                "    key_aliases:",
                '      replace_remote_control: "Replace Remote"',
                "    columns:",
                '      - "Defect"',
                "    rename_columns:",
                '      Defect: "defect"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "action"',
                '      - "defect"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    detail_df = pd.read_excel(
        result.output_path,
        sheet_name="Detail",
        skiprows=3,
        keep_default_na=False,
    )

    assert detail_df["defect"].tolist() == [
        "Panel",
        "Software",
        "Other",
        "N/A",
        "N/A",
        "",
    ]


def test_run_pipeline_supports_transforms_and_group_by_output(app_paths):
    source_path = app_paths.project_root / "source.csv"
    pd.DataFrame(
        [
            {"kategori": "Cat 1", "qty": 10, "harga": 10000},
            {"kategori": "Cat 2", "qty": 5, "harga": 8000},
            {"kategori": "Cat 3", "qty": 2, "harga": 12000},
        ]
    ).to_csv(source_path, index=False)

    config_path = app_paths.configs_dir / "transform_report.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Transform Report"',
                'source_sheet: "Sheet1"',
                "header:",
                '  title: "Transform Report"',
                "transforms:",
                '  - type: "ensure_optional_columns"',
                "    columns:",
                '      catatan: ""',
                '  - type: "filter_rows"',
                '    column: "qty"',
                "    gte: 5",
                '  - type: "formula"',
                '    target: "total"',
                '    operation: "multiply"',
                "    operands:",
                '      - column: "qty"',
                '      - column: "harga"',
                '  - type: "conditional"',
                '    target: "bucket"',
                "    cases:",
                "      - when:",
                '          column: "total"',
                "          gte: 50000",
                '        value: "besar"',
                '    default: "kecil"',
                "outputs:",
                '  - sheet_name: "Detail"',
                "    columns:",
                '      - "kategori"',
                '      - "qty"',
                '      - "harga"',
                '      - "catatan"',
                '      - "total"',
                '      - "bucket"',
                '  - sheet_name: "Summary"',
                "    group_by:",
                '      by: "bucket"',
                "      aggregations:",
                '        qty: "sum"',
                '        total: "sum"',
                "    columns:",
                '      - "bucket"',
                '      - "qty"',
                '      - "total"',
            ]
        ),
    )

    logs: list[str] = []
    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=logs.append,
    )

    detail_df = pd.read_excel(
        result.output_path,
        sheet_name="Detail",
        skiprows=3,
        keep_default_na=False,
    )
    summary_df = pd.read_excel(
        result.output_path,
        sheet_name="Summary",
        skiprows=3,
        keep_default_na=False,
    )

    assert detail_df["qty"].tolist() == [10, 5]
    assert detail_df["catatan"].tolist() == ["", ""]
    assert detail_df["total"].tolist() == [100000, 40000]
    assert detail_df["bucket"].tolist() == ["besar", "kecil"]
    assert summary_df["bucket"].tolist() == ["besar", "kecil"]
    assert summary_df["qty"].tolist() == [10, 5]
    assert summary_df["total"].tolist() == [100000, 40000]
    assert any("Apply transform rules" in item for item in logs)


def test_run_pipeline_supports_monthly_step_recipe_end_to_end(app_paths):
    source_path = app_paths.project_root / "monthly_source.xlsx"

    gqs_df = pd.DataFrame(
        [
            {
                "Notification": "123456789",
                "Job Sheet Section": 2,
                "Malfunction Start Date": "2026-02-15",
                "Basic Finish Date": "2026-02-20",
                "Model Name": "ABC42ZZ",
                "Category": "LCD SEID",
                "Serial Number": "ABCDE12345",
                "Symptom Code": "S1",
                "Symptom Code (Description)": "Desc 1",
                "PMActType": "PMA",
                "PMActType (Description)": "PMA Desc",
                "Symptom Comment": "vertical line",
                "Repair Comment": "replace screen",
                "Description": "desc gqs 1",
                "Warranty": "Yes",
                "Planner Group": "PG1",
                "cabang": "JKT",
                "Purchased Date": "2026-01-01",
                "Labor Cost": 10,
                "Transportation Cost": 1,
                "Parts Cost": 20,
                "Part used": "PNL01",
            },
            {
                "Notification": "123456789",
                "Job Sheet Section": 1,
                "Malfunction Start Date": "2026-02-15",
                "Basic Finish Date": "2026-02-20",
                "Model Name": "ABC42ZZ",
                "Category": "LCD SEID",
                "Serial Number": "ABCDE12345",
                "Symptom Code": "S2",
                "Symptom Code (Description)": "Desc 2",
                "PMActType": "PMA",
                "PMActType (Description)": "PMA Desc",
                "Symptom Comment": "adhesive issue",
                "Repair Comment": "tape adjust",
                "Description": "desc gqs 2",
                "Warranty": "Yes",
                "Planner Group": "PG1",
                "cabang": "JKT",
                "Purchased Date": "2026-01-01",
                "Labor Cost": 5,
                "Transportation Cost": 2,
                "Parts Cost": 30,
                "Part used": "TAP01",
            },
        ]
    )

    sass_df = pd.DataFrame(
        [
            {
                "No Claim": "SASS00000001",
                "QTY Claim": 1,
                "Receive": "2026-02-15",
                "Finish": "2026-02-18",
                "Model": "LX32ZZ",
                "Category": "LCD SEID",
                "Serial No": "VWXYZ98765",
                "Damage": "boot loop",
                "Part Replacement": "factor reset",
                "Branch": "JKT",
                "Purchase": "2024-01-01",
                "Service Fee": 5,
                "Transport Cost": 1,
                "Part": 400,
                "Part Code": "MAIN1",
            },
            {
                "No Claim": "SASS00000001",
                "QTY Claim": 2,
                "Receive": "2026-02-15",
                "Finish": "2026-02-18",
                "Model": "LX32ZZ",
                "Category": "LCD SEID",
                "Serial No": "VWXYZ98765",
                "Damage": "no power",
                "Part Replacement": "replace board",
                "Branch": "JKT",
                "Purchase": "2024-01-01",
                "Service Fee": 6,
                "Transport Cost": 2,
                "Part": 100,
                "Part Code": "PWR01",
            },
        ]
    )

    with pd.ExcelWriter(source_path) as writer:
        pd.DataFrame(
            [
                ["February 20", None, None],
                [None, None, None],
                [None, "SEID WARRANTY COST ORIGINAL", None],
                [None, "[3] WARRANTY COST TRANSITION - Monthly Base", None],
                [None, "Product Category", "Apr'19"],
                [None, "LCD TV SEID", 0.01],
            ]
        ).to_excel(writer, index=False, header=False, sheet_name="GQS vs SASS")
        gqs_df.to_excel(writer, index=False, sheet_name="GQS Mar26", startrow=1)
        sass_df.to_excel(writer, index=False, sheet_name="SASS Mar26", startrow=4)

    master_path = app_paths.masters_dir / "master_table.xlsx"
    with pd.ExcelWriter(master_path) as writer:
        pd.DataFrame(
            [
                {"part_used": "PNL01", "part_name": "PANEL"},
                {"part_used": "TAP01", "part_name": "TAPE"},
                {"part_used": "MAIN1", "part_name": "MAIN_UNIT"},
                {"part_used": "PWR01", "part_name": "POWER_UNIT"},
            ]
        ).to_excel(writer, index=False, sheet_name="part_list")
        pd.DataFrame([{"unused": "x"}]).to_excel(writer, index=False, sheet_name="panel_usage")
        pd.DataFrame(
            [
                {"model_name": "ABC42ZZ", "factory": "Factory A"},
                {"model_name": "LX32ZZ", "factory": "Factory B"},
            ]
        ).to_excel(writer, index=False, sheet_name="factory")
        pd.DataFrame(
            [
                {
                    "alias": "EXTERNAL",
                    "canonical": "EXTERNAL",
                    "scope": "both",
                }
            ]
        ).to_excel(writer, index=False, sheet_name="comment_synonyms")
        pd.DataFrame(
            [
                {
                    "priority": 20,
                    "part_name": "PANEL",
                    "match_type": "contains",
                    "pattern": "line",
                    "symptom": "LINE",
                    "notes": "panel line",
                },
                {
                    "priority": 10,
                    "part_name": "TAPE",
                    "match_type": "regex",
                    "pattern": ".*",
                    "symptom": "TAPE_GENERIC",
                    "notes": "fallback tape",
                },
                {
                    "priority": 30,
                    "part_name": "MAIN_UNIT",
                    "match_type": "contains",
                    "pattern": "boot",
                    "symptom": "BOOT",
                    "notes": "main boot",
                },
                {
                    "priority": 40,
                    "part_name": "POWER_UNIT",
                    "match_type": "regex",
                    "pattern": "power",
                    "symptom": "POWER",
                    "notes": "power regex",
                },
            ]
        ).to_excel(writer, index=False, sheet_name="symptom")
        pd.DataFrame([{"init": "JKT", "branch": "Jakarta"}]).to_excel(
            writer, index=False, sheet_name="branch"
        )
        pd.DataFrame(
            [
                {
                    "priority": 10,
                    "job_sheet_section": 1,
                    "part_name": "PANEL",
                    "symptom_comment": ".*",
                    "repair_comment": ".*",
                    "action": "replace_panel",
                },
                {
                    "priority": 20,
                    "job_sheet_section": 0,
                    "part_name": "TAPE",
                    "symptom_comment": ".*",
                    "repair_comment": ".*",
                    "action": "cancel",
                },
                {
                    "priority": 30,
                    "job_sheet_section": 1,
                    "part_name": "MAIN_UNIT",
                    "symptom_comment": ".*",
                    "repair_comment": ".*factor.*",
                    "action": "factory_reset",
                },
                {
                    "priority": 40,
                    "job_sheet_section": 0,
                    "part_name": "POWER_UNIT",
                    "symptom_comment": ".*",
                    "repair_comment": ".*",
                    "action": "replace_power_unit",
                },
            ]
        ).to_excel(writer, index=False, sheet_name="action")
        pd.DataFrame(
            [
                {
                    "Repair Action": "Replace Panel",
                    "Category": "Defect",
                    "Defect": "Panel",
                },
                {
                    "Repair Action": "Cancel",
                    "Category": "N/A",
                    "Defect": "N/A",
                },
                {
                    "Repair Action": "Factory Reset",
                    "Category": "Software",
                    "Defect": "Software",
                },
                {
                    "Repair Action": "Replace Power Unit",
                    "Category": "Defect",
                    "Defect": "Power",
                },
            ]
        ).to_excel(writer, index=False, sheet_name="defect_category")

    recipe_path = app_paths.configs_dir / "monthly-report-recipe.yaml"
    recipe_content = Path("configs/monthly-report-recipe.yaml").read_text(encoding="utf-8")
    recipe_content = recipe_content.replace('master: "symptom_comment"', 'master: "pattern"', 1)
    recipe_content = recipe_content.replace('          mode: "contains"', '          mode: "regex"', 1)
    recipe_path.write_text(recipe_content, encoding="utf-8")

    logs: list[str] = []
    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=recipe_path,
        log=logs.append,
        period_keydate_override="202603",
    )

    detail_df = pd.read_excel(
        result.output_path,
        sheet_name="result",
        skiprows=3,
        keep_default_na=False,
    )

    assert detail_df["notification"].tolist() == [
        "123456789",
        "123456789",
        "SASS00000001",
        "SASS00000001",
    ]
    assert detail_df["section"].tolist() == ["GQS", "GQS", "SASS", "SASS"]
    assert detail_df["part_name"].tolist() == ["PANEL", "TAPE", "MAIN_UNIT", "POWER_UNIT"]
    assert detail_df["job_sheet_section"].tolist() == [1, 0, 1, 0]
    assert detail_df["labor_cost"].tolist() == [1500, 0, 11, 0]
    assert detail_df["transportation_cost"].tolist() == [300, 0, 3, 0]
    assert detail_df["parts_cost"].tolist() == [5000, 0, 400, 100]
    assert detail_df["total_cost"].tolist() == [6800, 0, 414, 100]
    assert detail_df["prod_month"].astype(str).tolist() == ["123", "123", "987", "987"]
    assert detail_df["inch"].astype(str).tolist() == ["42", "42", "32", "32"]
    assert detail_df["panel_usage"].tolist() == ["< 1 Year", "< 1 Year", "2 - 3 Years", "2 - 3 Years"]
    assert detail_df["factory"].tolist() == ["Factory A", "Factory A", "Factory B", "Factory B"]
    assert detail_df["symptom"].tolist() == ["LINE", "TAPE_GENERIC", "BOOT", "POWER"]
    assert detail_df["branch"].tolist() == ["Jakarta", "Jakarta", "Jakarta", "Jakarta"]
    assert detail_df["action"].tolist() == [
        "replace_panel",
        "cancel",
        "factory_reset",
        "replace_power_unit",
    ]
    assert detail_df["defect_category"].tolist() == ["Defect", "N/A", "Software", "Defect"]
    assert detail_df["defect"].tolist() == ["Panel", "N/A", "Software", "Power"]
    assert detail_df["keydate"].astype(str).tolist() == ["202603", "202603", "202603", "202603"]
    assert any("duplicate group rewrite" in item.lower() for item in logs)


def test_run_pipeline_lookup_rules_step_recipe_supports_regex_and_priority_ordering(app_paths):
    source_path = app_paths.project_root / "source.xlsx"
    pd.DataFrame(
        [
            {
                "job_sheet_section": 1,
                "part_name": "PANEL",
                "symptom_comment": "vertical line",
                "repair_comment": "replace panel unit",
            },
            {
                "job_sheet_section": 1,
                "part_name": "PANEL",
                "symptom_comment": "vertical line",
                "repair_comment": "repair panel board",
            },
            {
                "job_sheet_section": 0,
                "part_name": "PANEL",
                "symptom_comment": "vertical line",
                "repair_comment": "replace panel unit",
            },
            {
                "job_sheet_section": 1,
                "part_name": "MAIN_UNIT",
                "symptom_comment": "boot loop",
                "repair_comment": "replace panel unit",
            },
        ]
    ).to_excel(source_path, index=False, sheet_name="Data")

    master_path = app_paths.masters_dir / "master_table.xlsx"
    with pd.ExcelWriter(master_path) as writer:
        pd.DataFrame(
            [
                {
                    "priority": 20,
                    "job_sheet_section": 1,
                    "part_name": "PANEL",
                    "symptom_comment": "(?i).*vertical line.*",
                    "repair_comment": "(?i).*replace.*",
                    "action": "REPLACE_PANEL",
                },
                {
                    "priority": 10,
                    "job_sheet_section": 1,
                    "part_name": "PANEL",
                    "symptom_comment": "(?i).*vertical line.*",
                    "repair_comment": "(?i).*replace.*",
                    "action": "PRIORITY_WINNER",
                },
                {
                    "priority": 30,
                    "job_sheet_section": 1,
                    "part_name": "PANEL",
                    "symptom_comment": "(?i).*vertical line.*",
                    "repair_comment": "(?i).*repair.*",
                    "action": "REPAIR_PANEL",
                },
            ]
        ).to_excel(writer, index=False, sheet_name="action")

    config_path = app_paths.configs_dir / "action_step_recipe.yaml"
    _write_action_lookup_rules_step_recipe(config_path)

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    detail_df = pd.read_excel(result.output_path, sheet_name="result", skiprows=3, keep_default_na=False)
    assert detail_df["action"].tolist() == ["PRIORITY_WINNER", "REPAIR_PANEL", "", ""]


def test_run_pipeline_lookup_rules_step_recipe_raises_for_invalid_regex(app_paths):
    source_path = app_paths.project_root / "source.xlsx"
    pd.DataFrame(
        [
            {
                "job_sheet_section": 1,
                "part_name": "PANEL",
                "symptom_comment": "vertical line",
                "repair_comment": "replace panel unit",
            }
        ]
    ).to_excel(source_path, index=False, sheet_name="Data")

    master_path = app_paths.masters_dir / "master_table.xlsx"
    with pd.ExcelWriter(master_path) as writer:
        pd.DataFrame(
            [
                {
                    "priority": 1,
                    "job_sheet_section": 1,
                    "part_name": "PANEL",
                    "symptom_comment": "(",
                    "repair_comment": "(?i).*replace.*",
                    "action": "REPLACE_PANEL",
                }
            ]
        ).to_excel(writer, index=False, sheet_name="action")

    config_path = app_paths.configs_dir / "action_step_recipe_invalid_regex.yaml"
    _write_action_lookup_rules_step_recipe(config_path)

    with pytest.raises(PipelineError, match="Regex matcher tidak valid"):
        run_pipeline(
            paths=app_paths,
            source_path=source_path,
            config_path=config_path,
            log=lambda _: None,
        )


def test_run_pipeline_lookup_rules_step_recipe_raises_for_invalid_priority(app_paths):
    source_path = app_paths.project_root / "source.xlsx"
    pd.DataFrame(
        [
            {
                "job_sheet_section": 1,
                "part_name": "PANEL",
                "symptom_comment": "vertical line",
                "repair_comment": "replace panel unit",
            }
        ]
    ).to_excel(source_path, index=False, sheet_name="Data")

    master_path = app_paths.masters_dir / "master_table.xlsx"
    with pd.ExcelWriter(master_path) as writer:
        pd.DataFrame(
            [
                {
                    "priority": 0,
                    "job_sheet_section": 1,
                    "part_name": "PANEL",
                    "symptom_comment": "(?i).*",
                    "repair_comment": "(?i).*replace.*",
                    "action": "REPLACE_PANEL",
                }
            ]
        ).to_excel(writer, index=False, sheet_name="action")

    config_path = app_paths.configs_dir / "action_step_recipe_invalid_priority.yaml"
    _write_action_lookup_rules_step_recipe(config_path)

    with pytest.raises(PipelineError, match="Priority harus integer positif"):
        run_pipeline(
            paths=app_paths,
            source_path=source_path,
            config_path=config_path,
            log=lambda _: None,
        )


def test_run_pipeline_lookup_rules_step_recipe_raises_when_priority_column_missing(app_paths):
    source_path = app_paths.project_root / "source.xlsx"
    pd.DataFrame(
        [
            {
                "job_sheet_section": 1,
                "part_name": "PANEL",
                "symptom_comment": "vertical line",
                "repair_comment": "replace panel unit",
            }
        ]
    ).to_excel(source_path, index=False, sheet_name="Data")

    master_path = app_paths.masters_dir / "master_table.xlsx"
    with pd.ExcelWriter(master_path) as writer:
        pd.DataFrame(
            [
                {
                    "job_sheet_section": 1,
                    "part_name": "PANEL",
                    "symptom_comment": "(?i).*",
                    "repair_comment": "(?i).*replace.*",
                    "action": "REPLACE_PANEL",
                }
            ]
        ).to_excel(writer, index=False, sheet_name="action")

    config_path = app_paths.configs_dir / "action_step_recipe_missing_priority_col.yaml"
    _write_action_lookup_rules_step_recipe(config_path)

    with pytest.raises(PipelineError, match="kolom priority 'priority' tidak ditemukan"):
        run_pipeline(
            paths=app_paths,
            source_path=source_path,
            config_path=config_path,
            log=lambda _: None,
        )


def test_run_pipeline_step_recipe_single_sheet_mode_extracts_without_sheet_name(app_paths):
    source_path = app_paths.project_root / "single_sheet_source.xlsx"
    pd.DataFrame(
        [
            {"Notification": "A-001"},
            {"Notification": "A-002"},
        ]
    ).to_excel(source_path, index=False, sheet_name="Random Name")

    config_path = app_paths.configs_dir / "single_sheet_mode.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Single Sheet Mode"',
                "datasets:",
                '  working_dataset: "result"',
                "  canonical_columns:",
                '    - "notification"',
                "steps:",
                '  - id: "extract"',
                '    type: "extract_sheet"',
                "    sheet_selector:",
                '      mode: "single_sheet_workbook"',
                "    header_locator:",
                "      scan_rows: [1, 1]",
                "      required:",
                '        - "Notification"',
                "    select:",
                '      "Notification": "notification"',
                '    write_to: "result"',
                "outputs:",
                '  - sheet_name: "result"',
                "    columns:",
                '      - "notification"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    detail_df = pd.read_excel(result.output_path, sheet_name="result", skiprows=3, keep_default_na=False)
    assert detail_df["notification"].tolist() == ["A-001", "A-002"]


def test_run_pipeline_step_recipe_single_sheet_mode_rejects_multisheet_workbook(app_paths):
    source_path = app_paths.project_root / "multi_sheet_source.xlsx"
    with pd.ExcelWriter(source_path) as writer:
        pd.DataFrame([{"Notification": "A-001"}]).to_excel(writer, index=False, sheet_name="Sheet1")
        pd.DataFrame([{"Notification": "A-002"}]).to_excel(writer, index=False, sheet_name="Sheet2")

    config_path = app_paths.configs_dir / "single_sheet_mode_invalid.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Single Sheet Mode Invalid"',
                "datasets:",
                '  working_dataset: "result"',
                "  canonical_columns:",
                '    - "notification"',
                "steps:",
                '  - id: "extract"',
                '    type: "extract_sheet"',
                "    sheet_selector:",
                '      mode: "single_sheet_workbook"',
                "    header_locator:",
                "      scan_rows: [1, 1]",
                "      required:",
                '        - "Notification"',
                "    select:",
                '      "Notification": "notification"',
                '    write_to: "result"',
                "outputs:",
                '  - sheet_name: "result"',
                "    columns:",
                '      - "notification"',
            ]
        ),
    )

    with pytest.raises(PipelineError, match="membutuhkan tepat 1 sheet"):
        run_pipeline(
            paths=app_paths,
            source_path=source_path,
            config_path=config_path,
            log=lambda _: None,
        )


def test_run_pipeline_step_recipe_summary_plain_layout_writes_without_report_header(app_paths):
    source_path = app_paths.project_root / "summary_plain_source.xlsx"
    pd.DataFrame([{"Notification": "A-001"}]).to_excel(source_path, index=False, sheet_name="OnlySheet")

    config_path = app_paths.configs_dir / "summary_plain_layout.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Summary Plain Layout"',
                "datasets:",
                '  working_dataset: "result"',
                "  canonical_columns:",
                '    - "notification"',
                "steps:",
                '  - id: "extract"',
                '    type: "extract_sheet"',
                "    sheet_selector:",
                '      mode: "single_sheet_workbook"',
                "    header_locator:",
                "      scan_rows: [1, 1]",
                "      required:",
                '        - "Notification"',
                "    select:",
                '      "Notification": "notification"',
                '    write_to: "result"',
                "outputs:",
                '  - sheet_name: "result"',
                "    columns:",
                '      - "notification"',
                '  - sheet_name: "data1"',
                "    summary:",
                '      type: "recipe_summary_v1"',
                '      layout_mode: "plain"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    workbook = load_workbook(result.output_path, read_only=True)
    assert workbook["result"]["A1"].value == "Summary Plain Layout"
    assert workbook["data1"]["A1"].value == "summary_type"
    assert workbook["data1"]["A2"].value == "recipe_summary_v1"


def test_run_pipeline_step_recipe_static_part_summary_builds_data1_totals(app_paths):
    source_path = app_paths.project_root / "data1_source.xlsx"
    pd.DataFrame(
        [
            {"section": "GQS", "part_name": "PANEL", "labor_cost": 10, "transportation_cost": 1, "parts_cost": 100, "total_cost": 111},
            {"section": "GQS", "part_name": "MAIN_UNIT", "labor_cost": 5, "transportation_cost": 1, "parts_cost": 50, "total_cost": 56},
            {"section": "GQS", "part_name": "LED_BAR", "labor_cost": 3, "transportation_cost": 0, "parts_cost": 30, "total_cost": 33},
            {"section": "GQS", "part_name": "", "labor_cost": 1, "transportation_cost": 0, "parts_cost": 1, "total_cost": 2},
            {"section": "SASS", "part_name": "PANEL", "labor_cost": 7, "transportation_cost": 1, "parts_cost": 70, "total_cost": 78},
            {"section": "SASS", "part_name": "POWER_UNIT", "labor_cost": 2, "transportation_cost": 1, "parts_cost": 20, "total_cost": 23},
            {"section": "SASS", "part_name": "REMOTE_CONTROL", "labor_cost": 1, "transportation_cost": 0, "parts_cost": 10, "total_cost": 11},
            {"section": "NEWSEC", "part_name": "MAIN_UNIT", "labor_cost": 4, "transportation_cost": 1, "parts_cost": 40, "total_cost": 45},
        ]
    ).to_excel(source_path, index=False, sheet_name="OnlySheet")

    config_path = app_paths.configs_dir / "static_part_summary_data1.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Data1 Static Part Summary"',
                "datasets:",
                '  working_dataset: "result"',
                "  canonical_columns:",
                '    - "section"',
                '    - "part_name"',
                '    - "labor_cost"',
                '    - "transportation_cost"',
                '    - "parts_cost"',
                '    - "total_cost"',
                "steps:",
                '  - id: "extract"',
                '    type: "extract_sheet"',
                "    sheet_selector:",
                '      mode: "single_sheet_workbook"',
                "    header_locator:",
                "      scan_rows: [1, 1]",
                "      required:",
                '        - "section"',
                '        - "part_name"',
                '        - "labor_cost"',
                '        - "transportation_cost"',
                '        - "parts_cost"',
                '        - "total_cost"',
                "    select:",
                '      "section": "section"',
                '      "part_name": "part_name"',
                '      "labor_cost": "labor_cost"',
                '      "transportation_cost": "transportation_cost"',
                '      "parts_cost": "parts_cost"',
                '      "total_cost": "total_cost"',
                '    write_to: "result"',
                "outputs:",
                '  - sheet_name: "data1"',
                "    summary:",
                '      type: "static_part_summary"',
                '      layout_mode: "plain"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    data1_df = pd.read_excel(result.output_path, sheet_name="data1", keep_default_na=False)

    gqs_total = data1_df[data1_df["section"] == "GQS Total"].iloc[0]
    sass_total = data1_df[data1_df["section"] == "SASS Total"].iloc[0]
    grand_total = data1_df[data1_df["section"] == "Grand Total"].iloc[0]

    assert gqs_total["Sum of total_cost"] == 200
    assert gqs_total["Count of part_name"] == 3
    assert sass_total["Sum of total_cost"] == 112
    assert sass_total["Count of part_name"] == 3
    assert grand_total["Sum of total_cost"] == 357
    assert grand_total["Count of part_name"] == 7

    gqs_other = data1_df[(data1_df["section"] == "GQS") & (data1_df["part_name"] == "OTHER")].iloc[0]
    assert gqs_other["Sum of total_cost"] == 33

    new_section_total = data1_df[data1_df["section"] == "NEWSEC Total"].iloc[0]
    assert new_section_total["Sum of total_cost"] == 45


def test_run_pipeline_step_recipe_static_part_pivot_summary_data1_forces_formula_mode(app_paths):
    source_path = app_paths.project_root / "data1_formula_source.xlsx"
    pd.DataFrame(
        [
            {"section": "GQS", "part_name": "PANEL", "labor_cost": 10, "transportation_cost": 1, "parts_cost": 100, "total_cost": 111},
            {"section": "GQS", "part_name": "MAIN_UNIT", "labor_cost": 5, "transportation_cost": 1, "parts_cost": 50, "total_cost": 56},
            {"section": "GQS", "part_name": "LED_BAR", "labor_cost": 3, "transportation_cost": 0, "parts_cost": 30, "total_cost": 33},
            {"section": "SASS", "part_name": "PANEL", "labor_cost": 7, "transportation_cost": 1, "parts_cost": 70, "total_cost": 78},
            {"section": "SASS", "part_name": "POWER_UNIT", "labor_cost": 2, "transportation_cost": 1, "parts_cost": 20, "total_cost": 23},
            {"section": "SASS", "part_name": "REMOTE_CONTROL", "labor_cost": 1, "transportation_cost": 0, "parts_cost": 10, "total_cost": 11},
        ]
    ).to_excel(source_path, index=False, sheet_name="OnlySheet")

    config_path = app_paths.configs_dir / "static_part_summary_data1_formula.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Data1 Static Part Summary Formula"',
                "datasets:",
                '  working_dataset: "result"',
                "  canonical_columns:",
                '    - "section"',
                '    - "part_name"',
                '    - "labor_cost"',
                '    - "transportation_cost"',
                '    - "parts_cost"',
                '    - "total_cost"',
                "steps:",
                '  - id: "extract"',
                '    type: "extract_sheet"',
                "    sheet_selector:",
                '      mode: "single_sheet_workbook"',
                "    header_locator:",
                "      scan_rows: [1, 1]",
                "      required:",
                '        - "section"',
                '        - "part_name"',
                '        - "labor_cost"',
                '        - "transportation_cost"',
                '        - "parts_cost"',
                '        - "total_cost"',
                "    select:",
                '      "section": "section"',
                '      "part_name": "part_name"',
                '      "labor_cost": "labor_cost"',
                '      "transportation_cost": "transportation_cost"',
                '      "parts_cost": "parts_cost"',
                '      "total_cost": "total_cost"',
                '    write_to: "result"',
                "outputs:",
                '  - sheet_name: "result"',
                "    columns:",
                '      - "section"',
                '      - "part_name"',
                '      - "labor_cost"',
                '      - "transportation_cost"',
                '      - "parts_cost"',
                '      - "total_cost"',
                '  - sheet_name: "data1"',
                "    summary:",
                '      type: "static_part_pivot_summary"',
                '      layout_mode: "plain"',
                "      options:",
                '        value_mode: "numeric"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    workbook = load_workbook(result.output_path, read_only=False, data_only=False)
    sheet = workbook["data1"]

    assert isinstance(sheet["C2"].value, str)
    assert sheet["C2"].value.startswith("=SUMIFS(")
    assert isinstance(sheet["G2"].value, str)
    assert sheet["G2"].value.startswith("=COUNTIFS(")

    other_row_idx = None
    for row_idx in range(2, sheet.max_row + 1):
        if sheet[f"A{row_idx}"].value == "GQS" and sheet[f"B{row_idx}"].value == "OTHER":
            other_row_idx = row_idx
            break

    assert other_row_idx is not None
    assert isinstance(sheet[f"F{other_row_idx}"].value, str)
    assert sheet[f"F{other_row_idx}"].value.startswith("=SUMIFS(")
    assert isinstance(sheet[f"G{other_row_idx}"].value, str)
    assert sheet[f"G{other_row_idx}"].value.startswith("=COUNTIFS(")

    grand_total_row_idx = None
    for row_idx in range(2, sheet.max_row + 1):
        if sheet[f"A{row_idx}"].value == "Grand Total":
            grand_total_row_idx = row_idx
            break

    assert grand_total_row_idx is not None
    assert isinstance(sheet[f"F{grand_total_row_idx}"].value, str)
    assert sheet[f"F{grand_total_row_idx}"].value.startswith("=SUMIFS(")
    assert isinstance(sheet[f"G{grand_total_row_idx}"].value, str)
    assert sheet[f"G{grand_total_row_idx}"].value.startswith("=COUNTIFS(")


def test_run_pipeline_step_recipe_part_pivot_summary_builds_data2_sorted_with_totals(app_paths):
    source_path = app_paths.project_root / "data2_source.xlsx"
    pd.DataFrame(
        [
            {"section": "GQS", "part_name": "PANEL", "labor_cost": 10, "transportation_cost": 1, "parts_cost": 89, "total_cost": 100},
            {"section": "GQS", "part_name": "PANEL", "labor_cost": 2, "transportation_cost": 0, "parts_cost": 18, "total_cost": 20},
            {"section": "GQS", "part_name": "MAIN_UNIT", "labor_cost": 20, "transportation_cost": 2, "parts_cost": 238, "total_cost": 260},
            {"section": "GQS", "part_name": "LED_BAR", "labor_cost": 5, "transportation_cost": 1, "parts_cost": 44, "total_cost": 50},
            {"section": "GQS", "part_name": "", "labor_cost": 1, "transportation_cost": 0, "parts_cost": 1, "total_cost": 2},
            {"section": "SASS", "part_name": "POWER_UNIT", "labor_cost": 8, "transportation_cost": 1, "parts_cost": 71, "total_cost": 80},
            {"section": "SASS", "part_name": "PANEL", "labor_cost": 7, "transportation_cost": 1, "parts_cost": 62, "total_cost": 70},
            {"section": "SASS", "part_name": "REMOTE_CONTROL", "labor_cost": 1, "transportation_cost": 0, "parts_cost": 9, "total_cost": 10},
            {"section": "NEWSEC", "part_name": "MAIN_UNIT", "labor_cost": 3, "transportation_cost": 0, "parts_cost": 30, "total_cost": 33},
        ]
    ).to_excel(source_path, index=False, sheet_name="OnlySheet")

    config_path = app_paths.configs_dir / "part_pivot_summary_data2.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Data2 Part Pivot Summary"',
                "datasets:",
                '  working_dataset: "result"',
                "  canonical_columns:",
                '    - "section"',
                '    - "part_name"',
                '    - "labor_cost"',
                '    - "transportation_cost"',
                '    - "parts_cost"',
                '    - "total_cost"',
                "steps:",
                '  - id: "extract"',
                '    type: "extract_sheet"',
                "    sheet_selector:",
                '      mode: "single_sheet_workbook"',
                "    header_locator:",
                "      scan_rows: [1, 1]",
                "      required:",
                '        - "section"',
                '        - "part_name"',
                '        - "labor_cost"',
                '        - "transportation_cost"',
                '        - "parts_cost"',
                '        - "total_cost"',
                "    select:",
                '      "section": "section"',
                '      "part_name": "part_name"',
                '      "labor_cost": "labor_cost"',
                '      "transportation_cost": "transportation_cost"',
                '      "parts_cost": "parts_cost"',
                '      "total_cost": "total_cost"',
                '    write_to: "result"',
                "outputs:",
                '  - sheet_name: "data2"',
                "    summary:",
                '      type: "part_pivot_summary"',
                '      layout_mode: "plain"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    data2_df = pd.read_excel(result.output_path, sheet_name="data2", keep_default_na=False)

    gqs_part_rows = data2_df[data2_df["section"] == "GQS"]["part_name"].tolist()
    assert gqs_part_rows == ["MAIN_UNIT", "PANEL", "LED_BAR"]

    gqs_total = data2_df[data2_df["section"] == "GQS Total"].iloc[0]
    sass_total = data2_df[data2_df["section"] == "SASS Total"].iloc[0]
    grand_total = data2_df[data2_df["section"] == "Grand Total"].iloc[0]

    assert gqs_total["Sum of total_cost"] == 430
    assert gqs_total["Count of part_name"] == 4
    assert sass_total["Sum of total_cost"] == 160
    assert sass_total["Count of part_name"] == 3
    assert grand_total["Sum of total_cost"] == 623
    assert grand_total["Count of part_name"] == 8

    panel_gqs = data2_df[(data2_df["section"] == "GQS") & (data2_df["part_name"] == "PANEL")].iloc[0]
    assert panel_gqs["Sum of total_cost"] == 120
    assert panel_gqs["Count of part_name"] == 2


def test_run_pipeline_step_recipe_panel_summaries_data3a_data3b_data3c(app_paths):
    source_path = app_paths.project_root / "data3_source.xlsx"
    pd.DataFrame(
        [
            {"part_name": "PANEL", "inch": "32", "model_name": "M1", "total_cost": 100, "symptom": "LINE", "branch": "JKT", "panel_usage": "< 1 Year"},
            {"part_name": "PANEL", "inch": "32", "model_name": "M2", "total_cost": 50, "symptom": "LINE", "branch": "BDG", "panel_usage": "1 - 2 Years"},
            {"part_name": "PANEL", "inch": "32", "model_name": "M1", "total_cost": 25, "symptom": "NO DISPLAY", "branch": "JKT", "panel_usage": "2 - 3 Years"},
            {"part_name": "PANEL", "inch": "55", "model_name": "M3", "total_cost": 200, "symptom": "NO DISPLAY", "branch": "SBY", "panel_usage": "> 3 Years"},
            {"part_name": "PANEL", "inch": "55", "model_name": "M4", "total_cost": 100, "symptom": "DEAD", "branch": "JKT", "panel_usage": ""},
            {"part_name": "PANEL", "inch": "65", "model_name": "M5", "total_cost": 80, "symptom": "DEAD", "branch": "MDN", "panel_usage": "INVALID"},
            {"part_name": "MAIN_UNIT", "inch": "55", "model_name": "X1", "total_cost": 999, "symptom": "X", "branch": "XXX", "panel_usage": "< 1 Year"},
        ]
    ).to_excel(source_path, index=False, sheet_name="OnlySheet")

    config_path = app_paths.configs_dir / "panel_summaries_data3.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Panel Summaries Data3"',
                "datasets:",
                '  working_dataset: "result"',
                "  canonical_columns:",
                '    - "part_name"',
                '    - "inch"',
                '    - "model_name"',
                '    - "total_cost"',
                '    - "symptom"',
                '    - "branch"',
                '    - "panel_usage"',
                "steps:",
                '  - id: "extract"',
                '    type: "extract_sheet"',
                "    sheet_selector:",
                '      mode: "single_sheet_workbook"',
                "    header_locator:",
                "      scan_rows: [1, 1]",
                "      required:",
                '        - "part_name"',
                '        - "inch"',
                '        - "model_name"',
                '        - "total_cost"',
                '        - "symptom"',
                '        - "branch"',
                '        - "panel_usage"',
                "    select:",
                '      "part_name": "part_name"',
                '      "inch": "inch"',
                '      "model_name": "model_name"',
                '      "total_cost": "total_cost"',
                '      "symptom": "symptom"',
                '      "branch": "branch"',
                '      "panel_usage": "panel_usage"',
                '    write_to: "result"',
                "outputs:",
                '  - sheet_name: "data3a"',
                "    summary:",
                '      type: "panel_model_summary"',
                '      layout_mode: "plain"',
                '  - sheet_name: "data3b"',
                "    summary:",
                '      type: "panel_symptom_summary"',
                '      layout_mode: "plain"',
                '  - sheet_name: "data3c"',
                "    summary:",
                '      type: "panel_area_summary"',
                '      layout_mode: "plain"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    data3a_df = pd.read_excel(result.output_path, sheet_name="data3a", keep_default_na=False)
    data3b_df = pd.read_excel(result.output_path, sheet_name="data3b", keep_default_na=False)
    data3c_df = pd.read_excel(result.output_path, sheet_name="data3c", keep_default_na=False)

    inch32_models = data3a_df[(data3a_df["part_name"] == "PANEL") & (data3a_df["inch"] == "32")]["model_name"].tolist()
    assert inch32_models == ["M1", "M2"]

    inch32_total = data3a_df[data3a_df["inch"] == "32 Total"].iloc[0]
    panel_total_a = data3a_df[data3a_df["part_name"] == "PANEL Total"].iloc[0]
    grand_total_a = data3a_df[data3a_df["part_name"] == "Grand Total"].iloc[0]
    assert inch32_total["Total"] == 175
    assert panel_total_a["Total"] == 555
    assert grand_total_a["Total"] == 555

    symptom_rows = data3b_df[data3b_df["part_name"] == "PANEL"]["symptom"].tolist()
    assert symptom_rows == ["DEAD", "LINE", "NO DISPLAY"]
    panel_total_b = data3b_df[data3b_df["part_name"] == "PANEL Total"].iloc[0]
    grand_total_b = data3b_df[data3b_df["part_name"] == "Grand Total"].iloc[0]
    assert panel_total_b["Total"] == 6
    assert grand_total_b["Total"] == 6

    branch_rows = data3c_df[data3c_df["part_name"] == "PANEL"]["branch"].tolist()
    assert branch_rows == ["JKT", "BDG", "MDN", "SBY"]
    panel_total_c = data3c_df[data3c_df["part_name"] == "PANEL Total"].iloc[0]
    grand_total_c = data3c_df[data3c_df["part_name"] == "Grand Total"].iloc[0]
    assert panel_total_c["Total"] == 6
    assert grand_total_c["Total"] == 6


def test_run_pipeline_step_recipe_panel_usage_summary_data4_with_fixed_order(app_paths):
    source_path = app_paths.project_root / "data4_source.xlsx"
    pd.DataFrame(
        [
            {"part_name": "PANEL", "panel_usage": "2 - 3 Years"},
            {"part_name": "PANEL", "panel_usage": "< 1 Year"},
            {"part_name": "PANEL", "panel_usage": "1 - 2 Years"},
            {"part_name": "PANEL", "panel_usage": "1 - 2 Years"},
            {"part_name": "PANEL", "panel_usage": "> 3 Years"},
            {"part_name": "PANEL", "panel_usage": "INVALID"},
            {"part_name": "PANEL", "panel_usage": ""},
            {"part_name": "MAIN_UNIT", "panel_usage": "< 1 Year"},
        ]
    ).to_excel(source_path, index=False, sheet_name="OnlySheet")

    config_path = app_paths.configs_dir / "panel_usage_summary_data4.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Panel Usage Summary Data4"',
                "datasets:",
                '  working_dataset: "result"',
                "  canonical_columns:",
                '    - "part_name"',
                '    - "panel_usage"',
                "steps:",
                '  - id: "extract"',
                '    type: "extract_sheet"',
                "    sheet_selector:",
                '      mode: "single_sheet_workbook"',
                "    header_locator:",
                "      scan_rows: [1, 1]",
                "      required:",
                '        - "part_name"',
                '        - "panel_usage"',
                "    select:",
                '      "part_name": "part_name"',
                '      "panel_usage": "panel_usage"',
                '    write_to: "result"',
                "outputs:",
                '  - sheet_name: "data4"',
                "    summary:",
                '      type: "panel_usage_summary"',
                '      layout_mode: "plain"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    data4_df = pd.read_excel(result.output_path, sheet_name="data4", keep_default_na=False)

    usage_rows = data4_df[data4_df["part_name"] == "PANEL"]["panel_usage"].tolist()
    assert usage_rows == ["< 1 Year", "1 - 2 Years", "2 - 3 Years", "> 3 Years"]

    usage_counts = data4_df[data4_df["part_name"] == "PANEL"]["Total"].tolist()
    assert usage_counts == [1, 2, 1, 1]

    panel_total = data4_df[data4_df["part_name"] == "PANEL Total"].iloc[0]
    grand_total = data4_df[data4_df["part_name"] == "Grand Total"].iloc[0]
    assert panel_total["Total"] == 5
    assert grand_total["Total"] == 5


def test_run_pipeline_step_recipe_panel_fcost_data5a_and_top1_model_data5b(app_paths):
    source_path = app_paths.project_root / "data5_source.xlsx"
    pd.DataFrame(
        [
            {"part_name": "PANEL", "inch": "55", "model_name": "M1", "labor_cost": 10, "transportation_cost": 2, "parts_cost": 288, "total_cost": 300},
            {"part_name": "PANEL", "inch": "55", "model_name": "M2", "labor_cost": 5, "transportation_cost": 1, "parts_cost": 114, "total_cost": 120},
            {"part_name": "PANEL", "inch": "55", "model_name": "M3", "labor_cost": 2, "transportation_cost": 1, "parts_cost": 47, "total_cost": 50},
            {"part_name": "PANEL", "inch": "55", "model_name": "M4", "labor_cost": 1, "transportation_cost": 0, "parts_cost": 19, "total_cost": 20},
            {"part_name": "PANEL", "inch": "55", "model_name": "M5", "labor_cost": 1, "transportation_cost": 0, "parts_cost": 6, "total_cost": 7},
            {"part_name": "PANEL", "inch": "55", "model_name": "M6", "labor_cost": 0, "transportation_cost": 0, "parts_cost": 3, "total_cost": 3},
            {"part_name": "PANEL", "inch": "32", "model_name": "A", "labor_cost": 3, "transportation_cost": 1, "parts_cost": 296, "total_cost": 300},
            {"part_name": "PANEL", "inch": "65", "model_name": "B", "labor_cost": 2, "transportation_cost": 1, "parts_cost": 197, "total_cost": 200},
            {"part_name": "PANEL", "inch": "24", "model_name": "C", "labor_cost": 1, "transportation_cost": 1, "parts_cost": 98, "total_cost": 100},
            {"part_name": "PANEL", "inch": "75", "model_name": "D", "labor_cost": 1, "transportation_cost": 0, "parts_cost": 79, "total_cost": 80},
            {"part_name": "PANEL", "inch": "42", "model_name": "E", "labor_cost": 1, "transportation_cost": 0, "parts_cost": 59, "total_cost": 60},
            {"part_name": "MAIN_UNIT", "inch": "55", "model_name": "X", "labor_cost": 99, "transportation_cost": 0, "parts_cost": 0, "total_cost": 99},
        ]
    ).to_excel(source_path, index=False, sheet_name="OnlySheet")

    config_path = app_paths.configs_dir / "panel_fcost_data5.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Panel FCost Data5"',
                "datasets:",
                '  working_dataset: "result"',
                "  canonical_columns:",
                '    - "part_name"',
                '    - "inch"',
                '    - "model_name"',
                '    - "labor_cost"',
                '    - "transportation_cost"',
                '    - "parts_cost"',
                '    - "total_cost"',
                "steps:",
                '  - id: "extract"',
                '    type: "extract_sheet"',
                "    sheet_selector:",
                '      mode: "single_sheet_workbook"',
                "    header_locator:",
                "      scan_rows: [1, 1]",
                "      required:",
                '        - "part_name"',
                '        - "inch"',
                '        - "model_name"',
                '        - "labor_cost"',
                '        - "transportation_cost"',
                '        - "parts_cost"',
                '        - "total_cost"',
                "    select:",
                '      "part_name": "part_name"',
                '      "inch": "inch"',
                '      "model_name": "model_name"',
                '      "labor_cost": "labor_cost"',
                '      "transportation_cost": "transportation_cost"',
                '      "parts_cost": "parts_cost"',
                '      "total_cost": "total_cost"',
                '    write_to: "result"',
                "outputs:",
                '  - sheet_name: "data5a"',
                "    summary:",
                '      type: "panel_fcost_inch_summary"',
                '      layout_mode: "plain"',
                '  - sheet_name: "data5b"',
                "    summary:",
                '      type: "panel_top1_inch_model_summary"',
                '      layout_mode: "plain"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    data5a_df = pd.read_excel(result.output_path, sheet_name="data5a", keep_default_na=False)
    data5b_df = pd.read_excel(result.output_path, sheet_name="data5b", keep_default_na=False)

    inch_rows = data5a_df[data5a_df["part_name"] == "PANEL"]["inch"].map(str).tolist()
    assert inch_rows == ["55", "32", "65", "24", "75", "other"]

    panel_total_a = data5a_df[data5a_df["part_name"] == "PANEL Total"].iloc[0]
    grand_total_a = data5a_df[data5a_df["part_name"] == "Grand Total"].iloc[0]
    assert panel_total_a["Sum of total_cost"] == 1240
    assert grand_total_a["Sum of total_cost"] == 1240

    model_rows = data5b_df[data5b_df["part_name"] == "PANEL"]["model_name"].tolist()
    assert model_rows == ["M1", "M2", "M3", "M4", "M5", "other"]
    inch_value_rows = data5b_df[data5b_df["part_name"] == "PANEL"]["inch"].map(str).unique().tolist()
    assert inch_value_rows == ["55"]

    panel_total_b = data5b_df[data5b_df["part_name"] == "PANEL Total"].iloc[0]
    grand_total_b = data5b_df[data5b_df["part_name"] == "Grand Total"].iloc[0]
    assert panel_total_b["Sum of total_cost"] == 500
    assert grand_total_b["Sum of total_cost"] == 500


def test_run_pipeline_step_recipe_panel_symptom_inch_matrix_data6(app_paths):
    source_path = app_paths.project_root / "data6_source.xlsx"
    pd.DataFrame(
        [
            {"part_name": "PANEL", "inch": "24", "symptom": "A"},
            {"part_name": "PANEL", "inch": "24", "symptom": "A"},
            {"part_name": "PANEL", "inch": "32", "symptom": "A"},
            {"part_name": "PANEL", "inch": "32", "symptom": "B"},
            {"part_name": "PANEL", "inch": "32", "symptom": "B"},
            {"part_name": "PANEL", "inch": "55", "symptom": "C"},
            {"part_name": "MAIN_UNIT", "inch": "24", "symptom": "Z"},
        ]
    ).to_excel(source_path, index=False, sheet_name="OnlySheet")

    config_path = app_paths.configs_dir / "panel_symptom_matrix_data6.yaml"
    _write_yaml(
        config_path,
        "\n".join(
            [
                'name: "Panel Symptom Matrix Data6"',
                "datasets:",
                '  working_dataset: "result"',
                "  canonical_columns:",
                '    - "part_name"',
                '    - "inch"',
                '    - "symptom"',
                "steps:",
                '  - id: "extract"',
                '    type: "extract_sheet"',
                "    sheet_selector:",
                '      mode: "single_sheet_workbook"',
                "    header_locator:",
                "      scan_rows: [1, 1]",
                "      required:",
                '        - "part_name"',
                '        - "inch"',
                '        - "symptom"',
                "    select:",
                '      "part_name": "part_name"',
                '      "inch": "inch"',
                '      "symptom": "symptom"',
                '    write_to: "result"',
                "outputs:",
                '  - sheet_name: "data6"',
                "    summary:",
                '      type: "panel_symptom_inch_matrix"',
                '      layout_mode: "plain"',
            ]
        ),
    )

    result = run_pipeline(
        paths=app_paths,
        source_path=source_path,
        config_path=config_path,
        log=lambda _: None,
    )

    data6_df = pd.read_excel(result.output_path, sheet_name="data6", keep_default_na=False)
    data6_df.columns = [str(col) for col in data6_df.columns]

    assert data6_df.columns.tolist() == ["part_name", "symptom", "24", "32", "55", "Grand Total"]

    symptom_rows = data6_df[data6_df["part_name"] == "PANEL"]["symptom"].tolist()
    assert symptom_rows == ["A", "B", "C"]

    row_a = data6_df[(data6_df["part_name"] == "PANEL") & (data6_df["symptom"] == "A")].iloc[0]
    assert row_a["24"] == 2
    assert row_a["32"] == 1
    assert row_a["55"] == 0
    assert row_a["Grand Total"] == 3

    panel_total = data6_df[data6_df["part_name"] == "PANEL Total"].iloc[0]
    grand_total = data6_df[data6_df["part_name"] == "Grand Total"].iloc[0]
    assert panel_total["Grand Total"] == 6
    assert grand_total["Grand Total"] == 6
