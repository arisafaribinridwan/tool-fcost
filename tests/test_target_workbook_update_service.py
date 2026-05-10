from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table

from app.services.target_workbook_update_service import (
    update_target_workbooks_by_model_series,
)


def _create_target_file(
    path: Path,
    headers: list[str],
    *,
    sheet_name: str = "raw",
    rows: list[list[object]] | None = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows or []:
        ws.append(row)
    wb.save(path)
    wb.close()


def test_update_target_appends_filtered_matching_rows(tmp_path):
    target_file = tmp_path / "X100.xlsx"
    _create_target_file(
        target_file,
        ["model_series", "notification", "part_used", "total_cost"],
        rows=[["old", "old", "old", "old"]],
    )

    data_df = pd.DataFrame(
        [
            {
                "model_series": "x100",
                "job_sheet_section": 1,
                "notification": "N1",
                "part_used": "P1",
                "total_cost": 10,
            },
            {
                "model_series": "x100",
                "job_sheet_section": 0,
                "notification": "N0",
                "part_used": "P0",
                "total_cost": 99,
            },
            {
                "model_series": "x100",
                "job_sheet_section": "1",
                "notification": "N2",
                "part_used": "P2",
                "total_cost": 20,
            },
        ]
    )

    results = update_target_workbooks_by_model_series(
        data_df=data_df,
        target_dir=tmp_path,
        match_column="model_series",
        target_sheet_name="raw",
        filter_column="job_sheet_section",
        filter_value=1,
    )

    assert len(results) == 1
    assert results[0].status == "updated"
    assert results[0].rows_written == 2

    wb = load_workbook(target_file)
    ws = wb["raw"]
    assert ws.cell(row=2, column=1).value == "old"
    assert ws.cell(row=2, column=2).value == "old"
    assert ws.cell(row=3, column=1).value == "x100"
    assert ws.cell(row=3, column=2).value == "N1"
    assert ws.cell(row=4, column=2).value == "N2"
    assert ws.max_row == 4
    wb.close()


def test_update_target_highlights_new_rows_and_clears_existing_data_fills(tmp_path):
    target_file = tmp_path / "X100.xlsx"
    _create_target_file(
        target_file,
        ["model_series", "notification", "part_used", "total_cost"],
        rows=[["old", "old", "old", "old"]],
    )

    wb = load_workbook(target_file)
    ws = wb["raw"]
    ws.cell(row=2, column=1).fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    wb.save(target_file)
    wb.close()

    data_df = pd.DataFrame(
        [
            {
                "model_series": "x100",
                "notification": "N1",
                "part_used": "P1",
                "total_cost": 10,
            },
            {
                "model_series": "x100",
                "notification": "N2",
                "part_used": "P2",
                "total_cost": 20,
            },
        ]
    )

    results = update_target_workbooks_by_model_series(
        data_df=data_df,
        target_dir=tmp_path,
        match_column="model_series",
        target_sheet_name="raw",
        new_row_color="E2EFDA",
    )

    assert results[0].status == "updated"
    assert results[0].rows_written == 2

    wb = load_workbook(target_file)
    ws = wb["raw"]
    assert ws.cell(row=2, column=1).fill.fill_type is None
    for row_index in (3, 4):
        for column_index in range(1, 5):
            cell_fill = ws.cell(row=row_index, column=column_index).fill
            assert cell_fill.fill_type == "solid"
            assert cell_fill.fgColor.rgb.endswith("E2EFDA")
    wb.close()


def test_update_target_creates_table_when_missing(tmp_path):
    target_file = tmp_path / "X100.xlsx"
    _create_target_file(target_file, ["model_series", "notification", "part_used"])

    data_df = pd.DataFrame(
        [{"model_series": "x100", "notification": "N1", "part_used": "P1"}]
    )

    results = update_target_workbooks_by_model_series(
        data_df=data_df,
        target_dir=tmp_path,
        match_column="model_series",
        target_sheet_name="raw",
        table_name="RawData",
        create_table_if_missing=True,
    )

    assert results[0].status == "updated"

    wb = load_workbook(target_file)
    ws = wb["raw"]
    assert "RawData" in ws.tables
    assert ws.tables["RawData"].ref == "A1:C2"
    wb.close()


def test_update_target_resizes_existing_table_after_append(tmp_path):
    target_file = tmp_path / "X100.xlsx"
    _create_target_file(
        target_file,
        ["model_series", "notification", "part_used"],
        rows=[["x100", "N0", "P0"]],
    )

    wb = load_workbook(target_file)
    ws = wb["raw"]
    ws.add_table(Table(displayName="RawData", ref="A1:C2"))
    wb.save(target_file)
    wb.close()

    data_df = pd.DataFrame(
        [
            {"model_series": "x100", "notification": "N1", "part_used": "P1"},
            {"model_series": "x100", "notification": "N2", "part_used": "P2"},
        ]
    )

    results = update_target_workbooks_by_model_series(
        data_df=data_df,
        target_dir=tmp_path,
        match_column="model_series",
        target_sheet_name="raw",
        table_name="RawData",
        create_table_if_missing=True,
    )

    assert results[0].status == "updated"
    assert results[0].rows_written == 2

    wb = load_workbook(target_file)
    ws = wb["raw"]
    assert ws.tables["RawData"].ref == "A1:C4"
    wb.close()


def test_update_target_skips_table_create_when_disabled(tmp_path):
    target_file = tmp_path / "X100.xlsx"
    _create_target_file(target_file, ["model_series", "notification"])

    data_df = pd.DataFrame([{"model_series": "x100", "notification": "N1"}])

    results = update_target_workbooks_by_model_series(
        data_df=data_df,
        target_dir=tmp_path,
        match_column="model_series",
        target_sheet_name="raw",
        table_name="RawData",
    )

    assert results[0].status == "updated"

    wb = load_workbook(target_file)
    ws = wb["raw"]
    assert "RawData" not in ws.tables
    wb.close()


def test_update_target_skips_existing_duplicate_key_rows(tmp_path):
    target_file = tmp_path / "X100.xlsx"
    _create_target_file(
        target_file,
        ["model_series", "notification", "part_used", "total_cost"],
        rows=[["x100", "N1", "P1", 10]],
    )

    data_df = pd.DataFrame(
        [
            {
                "model_series": "x100",
                "notification": "N1",
                "part_used": "P1",
                "total_cost": 999,
            },
            {
                "model_series": "x100",
                "notification": "N2",
                "part_used": "P2",
                "total_cost": 20,
            },
        ]
    )

    results = update_target_workbooks_by_model_series(
        data_df=data_df,
        target_dir=tmp_path,
        match_column="model_series",
        target_sheet_name="raw",
        duplicate_key_columns=("notification", "model_series", "part_used"),
    )

    assert results[0].status == "updated"
    assert results[0].rows_written == 1

    wb = load_workbook(target_file)
    ws = wb["raw"]
    assert ws.max_row == 3
    assert ws.cell(row=2, column=2).value == "N1"
    assert ws.cell(row=2, column=4).value == 10
    assert ws.cell(row=3, column=2).value == "N2"
    assert ws.cell(row=3, column=4).value == 20
    wb.close()


def test_update_target_skips_file_when_all_rows_already_exist(tmp_path):
    target_file = tmp_path / "X100.xlsx"
    _create_target_file(
        target_file,
        ["model_series", "notification", "part_used"],
        rows=[["x100", "N1", "P1"]],
    )

    data_df = pd.DataFrame(
        [{"model_series": "x100", "notification": "N1", "part_used": "P1"}]
    )

    results = update_target_workbooks_by_model_series(
        data_df=data_df,
        target_dir=tmp_path,
        match_column="model_series",
        target_sheet_name="raw",
        duplicate_key_columns=("notification", "model_series", "part_used"),
    )

    assert results[0].status == "skipped"
    assert results[0].rows_written == 0
    assert "sudah ada" in results[0].reason

    wb = load_workbook(target_file)
    ws = wb["raw"]
    assert ws.max_row == 2
    wb.close()


def test_update_target_fails_when_target_missing_duplicate_key_column(tmp_path):
    target_file = tmp_path / "X100.xlsx"
    _create_target_file(target_file, ["model_series", "notification"])

    data_df = pd.DataFrame(
        [{"model_series": "x100", "notification": "N1", "part_used": "P1"}]
    )

    results = update_target_workbooks_by_model_series(
        data_df=data_df,
        target_dir=tmp_path,
        match_column="model_series",
        target_sheet_name="raw",
        duplicate_key_columns=("notification", "model_series", "part_used"),
    )

    assert results[0].status == "failed"
    assert "part_used" in results[0].reason


def test_update_target_matches_files_with_order_prefix_when_enabled(tmp_path):
    target_file = tmp_path / "1-X100.xlsx"
    _create_target_file(target_file, ["model_series", "notification"])

    data_df = pd.DataFrame([{"model_series": "x100", "notification": "N1"}])

    results = update_target_workbooks_by_model_series(
        data_df=data_df,
        target_dir=tmp_path,
        match_column="model_series",
        target_sheet_name="raw",
        strip_order_prefix=True,
    )

    assert results[0].status == "updated"
    assert results[0].model_series_key == "x100"
    assert results[0].rows_written == 1

    wb = load_workbook(target_file)
    ws = wb["raw"]
    assert ws.cell(row=2, column=1).value == "x100"
    assert ws.cell(row=2, column=2).value == "N1"
    wb.close()


def test_update_target_processes_order_prefixed_files_by_number(tmp_path):
    for name in ["10-X100.xlsx", "2-X200.xlsx", "1-X300.xlsx"]:
        _create_target_file(tmp_path / name, ["model_series", "notification"])

    data_df = pd.DataFrame(
        [
            {"model_series": "x100", "notification": "N10"},
            {"model_series": "x200", "notification": "N2"},
            {"model_series": "x300", "notification": "N1"},
        ]
    )

    results = update_target_workbooks_by_model_series(
        data_df=data_df,
        target_dir=tmp_path,
        match_column="model_series",
        target_sheet_name="raw",
        strip_order_prefix=True,
    )

    assert [result.file_name for result in results] == [
        "1-X300.xlsx",
        "2-X200.xlsx",
        "10-X100.xlsx",
    ]
    assert all(result.status == "updated" for result in results)


def test_update_target_keeps_order_prefix_literal_when_disabled(tmp_path):
    _create_target_file(tmp_path / "1-X100.xlsx", ["model_series", "notification"])

    data_df = pd.DataFrame([{"model_series": "x100", "notification": "N1"}])

    results = update_target_workbooks_by_model_series(
        data_df=data_df,
        target_dir=tmp_path,
        match_column="model_series",
        target_sheet_name="raw",
    )

    assert results[0].status == "skipped"
    assert results[0].model_series_key == "1-x100"


def test_update_target_skips_when_no_match(tmp_path):
    target_file = tmp_path / "X200.xlsx"
    _create_target_file(target_file, ["model_series", "notification"])

    data_df = pd.DataFrame([{"model_series": "x100", "notification": "N1"}])

    results = update_target_workbooks_by_model_series(
        data_df=data_df,
        target_dir=tmp_path,
        match_column="model_series",
        target_sheet_name="raw",
    )

    assert results[0].status == "skipped"
    assert "Tidak ada data" in results[0].reason


def test_update_target_raises_when_no_xlsx(tmp_path):
    data_df = pd.DataFrame([{"model_series": "x100"}])

    with pytest.raises(ValueError, match="tidak memiliki file .xlsx"):
        update_target_workbooks_by_model_series(
            data_df=data_df,
            target_dir=tmp_path,
            match_column="model_series",
            target_sheet_name="raw",
        )
