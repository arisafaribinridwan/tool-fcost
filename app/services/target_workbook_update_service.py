from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class TargetFileUpdateResult:
    file_name: str
    model_series_key: str
    status: str
    rows_written: int
    reason: str = ""


def _normalize_key(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text.casefold()


_ORDER_PREFIX_PATTERN = re.compile(r"^\s*(\d+)\s*[-_]\s*(.+?)\s*$")


def _split_order_prefix(stem: str) -> tuple[int | None, str]:
    match = _ORDER_PREFIX_PATTERN.match(stem)
    if match is None:
        return None, stem
    return int(match.group(1)), match.group(2).strip()


def _get_target_excel_files(
    target_dir: Path,
    *,
    strip_order_prefix: bool = False,
) -> list[Path]:
    target_files = [
        path
        for path in target_dir.iterdir()
        if path.is_file() and path.suffix.lower() == ".xlsx"
    ]
    if not strip_order_prefix:
        return sorted(target_files, key=lambda item: item.name.casefold())

    def sort_key(path: Path) -> tuple[int, int, str]:
        order_number, model_stem = _split_order_prefix(path.stem)
        if order_number is None:
            return (1, 0, path.name.casefold())
        return (0, order_number, model_stem.casefold())

    return sorted(target_files, key=sort_key)


def _get_target_model_series_key(
    target_file: Path,
    *,
    strip_order_prefix: bool,
) -> str:
    stem = target_file.stem
    if strip_order_prefix:
        _, stem = _split_order_prefix(stem)
    return _normalize_key(stem)


def _normalize_filter_value(value: object) -> str:
    return _normalize_key(value)


def _build_row_key(
    row_values: dict[str, object],
    key_columns: tuple[str, ...],
) -> tuple[str, ...]:
    return tuple(_normalize_key(row_values.get(column)) for column in key_columns)


def _collect_existing_keys(
    worksheet: Worksheet,
    *,
    target_columns: list[str],
    duplicate_key_columns: tuple[str, ...],
) -> set[tuple[str, ...]]:
    column_index_by_name = {
        column: index + 1 for index, column in enumerate(target_columns)
    }
    missing_target_columns = [
        column for column in duplicate_key_columns if column not in column_index_by_name
    ]
    if missing_target_columns:
        raise ValueError(
            "Kolom duplicate key tidak ditemukan pada sheet target: "
            + ", ".join(missing_target_columns)
        )

    existing_keys: set[tuple[str, ...]] = set()
    for row_index in range(2, worksheet.max_row + 1):
        row_values = {
            column: worksheet.cell(
                row=row_index,
                column=column_index_by_name[column],
            ).value
            for column in duplicate_key_columns
        }
        row_key = _build_row_key(row_values, duplicate_key_columns)
        if any(row_key):
            existing_keys.add(row_key)
    return existing_keys


def _filter_new_rows(
    matched_df: pd.DataFrame,
    *,
    existing_keys: set[tuple[str, ...]],
    duplicate_key_columns: tuple[str, ...],
) -> pd.DataFrame:
    if not duplicate_key_columns:
        return matched_df

    missing_source_columns = [
        column for column in duplicate_key_columns if column not in matched_df.columns
    ]
    if missing_source_columns:
        raise ValueError(
            "Kolom duplicate key tidak ditemukan pada data source: "
            + ", ".join(missing_source_columns)
        )

    new_row_indexes: list[object] = []
    seen_in_batch: set[tuple[str, ...]] = set()
    for row_index, row in matched_df.iterrows():
        row_key = _build_row_key(row.to_dict(), duplicate_key_columns)
        if row_key in existing_keys or row_key in seen_in_batch:
            continue
        seen_in_batch.add(row_key)
        new_row_indexes.append(row_index)

    return matched_df.loc[new_row_indexes].copy()


_NO_FILL = PatternFill(fill_type=None)


def _clear_data_row_fills(worksheet: Worksheet) -> None:
    if worksheet.max_row < 2:
        return
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.fill = _NO_FILL


def _apply_row_fill(
    worksheet: Worksheet,
    *,
    row_index: int,
    column_count: int,
    color: str,
) -> None:
    fill = PatternFill(fill_type="solid", fgColor=color.upper())
    for column_index in range(1, column_count + 1):
        worksheet.cell(row=row_index, column=column_index).fill = fill


def _get_worksheet_table_by_name(worksheet: Worksheet, table_name: str) -> Table | None:
    table = worksheet.tables.get(table_name)
    return table if isinstance(table, Table) else None


def _workbook_has_table_name_outside_sheet(
    worksheet: Worksheet,
    *,
    table_name: str,
) -> bool:
    workbook = worksheet.parent
    for other_worksheet in workbook.worksheets:
        if other_worksheet is worksheet:
            continue
        if table_name in other_worksheet.tables:
            return True
    return False


def _ensure_excel_table(
    worksheet: Worksheet,
    *,
    table_name: str,
    column_count: int,
    create_if_missing: bool,
) -> None:
    if not table_name:
        return
    if column_count < 1 or worksheet.max_row < 2:
        return

    last_column_letter = get_column_letter(column_count)
    table_ref = f"A1:{last_column_letter}{worksheet.max_row}"
    table = _get_worksheet_table_by_name(worksheet, table_name)
    if table is not None:
        table.ref = table_ref
        return

    if not create_if_missing:
        return

    if _workbook_has_table_name_outside_sheet(worksheet, table_name=table_name):
        raise ValueError(
            f"Nama table '{table_name}' sudah dipakai di sheet lain pada workbook target."
        )

    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def update_target_workbooks_by_model_series(
    *,
    data_df: pd.DataFrame,
    target_dir: Path,
    match_column: str,
    target_sheet_name: str,
    filter_column: str | None = None,
    filter_value: object | None = None,
    duplicate_key_columns: tuple[str, ...] = (),
    new_row_color: str | None = None,
    strip_order_prefix: bool = False,
    table_name: str | None = None,
    create_table_if_missing: bool = False,
) -> list[TargetFileUpdateResult]:
    if not target_dir.exists() or not target_dir.is_dir():
        raise ValueError("Folder tujuan tidak ditemukan atau bukan folder.")

    if match_column not in data_df.columns:
        raise ValueError(
            f"Kolom match tidak ditemukan pada data source: {match_column}"
        )

    if filter_column is not None and filter_column not in data_df.columns:
        raise ValueError(
            f"Kolom filter tidak ditemukan pada data source: {filter_column}"
        )

    if filter_column is not None:
        expected_value = _normalize_filter_value(filter_value)
        filter_mask = (
            data_df[filter_column].map(_normalize_filter_value) == expected_value
        )
        data_df = data_df.loc[filter_mask].copy()

    target_files = _get_target_excel_files(
        target_dir,
        strip_order_prefix=strip_order_prefix,
    )
    if not target_files:
        raise ValueError("Folder tujuan tidak memiliki file .xlsx untuk diproses.")

    results: list[TargetFileUpdateResult] = []
    normalized_series = data_df[match_column].map(_normalize_key)

    for target_file in target_files:
        key = _get_target_model_series_key(
            target_file,
            strip_order_prefix=strip_order_prefix,
        )
        if not key:
            results.append(
                TargetFileUpdateResult(
                    file_name=target_file.name,
                    model_series_key=key,
                    status="skipped",
                    rows_written=0,
                    reason="Nama file kosong setelah normalisasi.",
                )
            )
            continue

        matched_df = data_df.loc[normalized_series == key].copy()
        if matched_df.empty:
            results.append(
                TargetFileUpdateResult(
                    file_name=target_file.name,
                    model_series_key=key,
                    status="skipped",
                    rows_written=0,
                    reason="Tidak ada data model_series yang cocok.",
                )
            )
            continue

        try:
            workbook = load_workbook(target_file)
            if target_sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{target_sheet_name}' tidak ditemukan.")
            worksheet = workbook[target_sheet_name]

            header_values = [cell.value for cell in worksheet[1]]
            target_columns = [
                str(value).strip()
                for value in header_values
                if isinstance(value, str) and value.strip()
            ]
            if not target_columns:
                raise ValueError("Header kolom pada baris 1 kosong.")

            write_columns = [
                column for column in target_columns if column in matched_df.columns
            ]
            if not write_columns:
                raise ValueError(
                    "Tidak ada kolom target yang cocok dengan data source."
                )

            if duplicate_key_columns:
                existing_keys = _collect_existing_keys(
                    worksheet,
                    target_columns=target_columns,
                    duplicate_key_columns=duplicate_key_columns,
                )
                matched_df = _filter_new_rows(
                    matched_df,
                    existing_keys=existing_keys,
                    duplicate_key_columns=duplicate_key_columns,
                )
                if matched_df.empty:
                    if table_name:
                        _ensure_excel_table(
                            worksheet,
                            table_name=table_name,
                            column_count=len(target_columns),
                            create_if_missing=create_table_if_missing,
                        )
                        workbook.save(target_file)
                    workbook.close()
                    results.append(
                        TargetFileUpdateResult(
                            file_name=target_file.name,
                            model_series_key=key,
                            status="skipped",
                            rows_written=0,
                            reason="Semua data sudah ada di sheet target.",
                        )
                    )
                    continue

            if new_row_color:
                _clear_data_row_fills(worksheet)

            for _, row in matched_df.iterrows():
                row_values: list[object] = []
                for target_column in target_columns:
                    if target_column in matched_df.columns:
                        value = row.get(target_column)
                        row_values.append(None if pd.isna(value) else value)
                    else:
                        row_values.append(None)
                worksheet.append(row_values)
                if new_row_color:
                    _apply_row_fill(
                        worksheet,
                        row_index=worksheet.max_row,
                        column_count=len(target_columns),
                        color=new_row_color,
                    )

            if table_name:
                _ensure_excel_table(
                    worksheet,
                    table_name=table_name,
                    column_count=len(target_columns),
                    create_if_missing=create_table_if_missing,
                )

            workbook.save(target_file)
            workbook.close()
            results.append(
                TargetFileUpdateResult(
                    file_name=target_file.name,
                    model_series_key=key,
                    status="updated",
                    rows_written=len(matched_df),
                )
            )
        except Exception as exc:
            results.append(
                TargetFileUpdateResult(
                    file_name=target_file.name,
                    model_series_key=key,
                    status="failed",
                    rows_written=0,
                    reason=str(exc),
                )
            )

    return results
