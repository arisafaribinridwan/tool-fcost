from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from app.utils.path_safety import resolve_runtime_relative_path


INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def sanitize_sheet_name(raw_name: str, used_names: set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("_", raw_name).strip()
    if not cleaned:
        cleaned = "Sheet"
    cleaned = cleaned[:31]

    candidate = cleaned
    suffix = 2
    while candidate in used_names:
        marker = f"_{suffix}"
        candidate = f"{cleaned[: 31 - len(marker)]}{marker}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def _build_period_text(source_df: pd.DataFrame, header_cfg: dict) -> str:
    period_col = header_cfg.get("period_from_column")
    if not isinstance(period_col, str) or period_col not in source_df.columns:
        return "Periode: -"

    parsed = pd.to_datetime(source_df[period_col], errors="coerce")
    parsed = parsed.dropna()
    if parsed.empty:
        return "Periode: -"

    start = parsed.min().strftime("%d/%m/%Y")
    end = parsed.max().strftime("%d/%m/%Y")
    return f"Periode: {start} - {end}"


def _apply_worksheet_style(
    worksheet,
    frame: pd.DataFrame,
    startrow: int,
    styling_cfg: dict,
    freeze_pane_default: str,
    row_types: list[str] | None = None,
) -> None:
    header_row = startrow + 1
    data_first_row = header_row + 1
    data_last_row = header_row + len(frame)
    max_col = max(1, len(frame.columns))

    header_color = str(styling_cfg.get("header_color", "4472C4")).upper()
    header_fill = PatternFill(fill_type="solid", fgColor=header_color)
    header_font = Font(
        name=str(styling_cfg.get("font", "Calibri")),
        bold=True,
        color="FFFFFF",
    )
    default_font = Font(name=str(styling_cfg.get("font", "Calibri")))
    bold_font = Font(name=str(styling_cfg.get("font", "Calibri")), bold=True)
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Row-type fill definitions
    zebra_enabled = bool(styling_cfg.get("zebra_stripe", False))
    zebra_fill = PatternFill(fill_type="solid", fgColor=str(styling_cfg.get("zebra_color", "F2F2F2")).upper())
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    subtotal_fill = PatternFill(fill_type="solid", fgColor=str(styling_cfg.get("subtotal_color", "FFF2CC")).upper())
    grand_total_color = str(styling_cfg.get("grand_total_color", "1F3864")).upper()
    grand_total_fill = PatternFill(fill_type="solid", fgColor=grand_total_color)
    grand_total_font = Font(name=str(styling_cfg.get("font", "Calibri")), bold=True, color="FFFFFF")

    for col_idx in range(1, max_col + 1):
        header_cell = worksheet.cell(row=header_row, column=col_idx)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.border = border

        if len(frame.columns) >= col_idx:
            column_name = frame.columns[col_idx - 1]
            sample_size = min(100, len(frame))
            # Skip formula strings when computing column width
            non_formula_lengths = [
                len(str(v))
                for v in frame[column_name].head(sample_size)
                if not (isinstance(v, str) and str(v).startswith("="))
            ]
            max_len = max(
                [len(str(column_name))]
                + (non_formula_lengths if non_formula_lengths else [10])
            )
            worksheet.column_dimensions[header_cell.column_letter].width = min(
                max(10, max_len + 2),
                48,
            )

    if len(frame) == 0:
        worksheet.freeze_panes = styling_cfg.get("freeze_pane", freeze_pane_default)
        return

    date_format = str(styling_cfg.get("date_format", "DD/MM/YYYY"))
    number_format = str(styling_cfg.get("number_format", "#,##0"))

    # Apply per-row styling
    data_row_counter = 0
    for row_idx in range(data_first_row, data_last_row + 1):
        frame_row_idx = row_idx - data_first_row
        row_type = row_types[frame_row_idx] if row_types and frame_row_idx < len(row_types) else "data"

        if row_type == "grand_total":
            row_fill = grand_total_fill
            row_font = grand_total_font
        elif row_type == "subtotal":
            row_fill = subtotal_fill
            row_font = bold_font
        else:
            row_fill = zebra_fill if (zebra_enabled and data_row_counter % 2 == 1) else white_fill
            row_font = default_font
            data_row_counter += 1

        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.font = row_font
            cell.fill = row_fill

    for col_idx, column_name in enumerate(frame.columns, start=1):
        series = frame[column_name]
        if pd.api.types.is_datetime64_any_dtype(series):
            for row_idx in range(data_first_row, data_last_row + 1):
                worksheet.cell(row=row_idx, column=col_idx).number_format = date_format
        elif pd.api.types.is_numeric_dtype(series):
            for row_idx in range(data_first_row, data_last_row + 1):
                worksheet.cell(row=row_idx, column=col_idx).number_format = number_format

    worksheet.freeze_panes = styling_cfg.get("freeze_pane", freeze_pane_default)


def _is_total_text(value: object) -> bool:
    text = str(value).strip().casefold()
    return bool(text) and (text == "grand total" or text.endswith(" total"))


def _fill_from_config(styling_cfg: dict, key: str, default: str) -> PatternFill:
    color = str(styling_cfg.get(key, default)).upper()
    return PatternFill(fill_type="solid", fgColor=color)


def _apply_summary_sheet_style(
    worksheet,
    frame: pd.DataFrame,
    startrow: int,
    styling_cfg: dict,
    row_types: list[str] | None,
    sheet_options: dict | None,
) -> None:
    """Apply presentation-style formatting for summary sheets only."""
    cfg = sheet_options or {}
    font_name = str(styling_cfg.get("font", "Calibri"))
    header_row = startrow + 1
    data_first_row = header_row + 1
    data_last_row = header_row + len(frame)
    max_col = max(1, len(frame.columns))

    title = str(cfg.get("title", "")).strip()
    subtitle = str(cfg.get("subtitle", "")).strip()

    if title:
        worksheet["A1"] = title
    if subtitle:
        worksheet["A2"] = subtitle
    if max_col > 1:
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=max_col,
        )
        worksheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=max_col,
        )

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = str(cfg.get("freeze_pane", "B5"))

    column_width = float(cfg.get("column_width", 13.0))
    for col_idx in range(1, max_col + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = column_width

    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    title_fill = _fill_from_config(styling_cfg, "summary_title_color", "1F4E78")
    section_total_header_fill = _fill_from_config(
        styling_cfg,
        "summary_header_section_total_color",
        "E4DFEC",
    )
    header_fills = [
        section_total_header_fill,
        _fill_from_config(styling_cfg, "summary_header_part_color", "BDD7EE"),
        _fill_from_config(styling_cfg, "summary_header_labor_color", "C6EFCE"),
        _fill_from_config(styling_cfg, "summary_header_transportation_color", "FFE699"),
        _fill_from_config(styling_cfg, "summary_header_parts_color", "FCE4D6"),
        section_total_header_fill,
        _fill_from_config(styling_cfg, "summary_header_count_color", "DDEBF7"),
    ]
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    zebra_fill = _fill_from_config(styling_cfg, "zebra_color", "F2F2F2")
    subtotal_fill = _fill_from_config(styling_cfg, "subtotal_color", "FFF2CC")
    grand_total_fill = _fill_from_config(styling_cfg, "summary_grand_total_color", "C6EFCE")

    worksheet["A1"].font = Font(name=font_name, bold=True, size=14, color="FFFFFF")
    worksheet["A1"].fill = title_fill
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A2"].font = Font(name=font_name, italic=True, size=10)
    worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, max_col + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.font = Font(name=font_name, bold=True, size=10)
        header_text = str(cell.value).strip().casefold()
        if header_text in {"section", "total"}:
            cell.fill = section_total_header_fill
        else:
            cell.fill = header_fills[(col_idx - 1) % len(header_fills)]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    data_row_counter = 0
    for row_idx in range(data_first_row, data_last_row + 1):
        frame_row_idx = row_idx - data_first_row
        if row_types and frame_row_idx < len(row_types):
            row_type = row_types[frame_row_idx]
        else:
            row_type = ""
        first_value = worksheet.cell(row=row_idx, column=1).value
        if not row_type:
            if str(first_value).strip().casefold() == "grand total":
                row_type = "grand_total"
            elif _is_total_text(first_value):
                row_type = "subtotal"
            else:
                row_type = "data"

        if row_type == "grand_total":
            row_fill = grand_total_fill
            row_font = Font(name=font_name, bold=True, size=9)
        elif row_type == "subtotal":
            row_fill = subtotal_fill
            row_font = Font(name=font_name, bold=True, size=9)
        else:
            row_fill = zebra_fill if data_row_counter % 2 == 1 else white_fill
            row_font = Font(name=font_name, size=9)
            data_row_counter += 1

        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = row_font
            cell.fill = row_fill
            cell.border = border
            vertical = "top" if col_idx == 1 and row_type == "data" else "center"
            horizontal = "right" if col_idx >= 3 else "left"
            if row_type in {"subtotal", "grand_total"} and col_idx == 1:
                horizontal = "center"
            cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
            if col_idx >= 3:
                cell.number_format = str(styling_cfg.get("number_format", "#,##0"))

    if data_last_row >= data_first_row:
        merge_start: int | None = None
        current_value = ""
        for row_idx in range(data_first_row, data_last_row + 2):
            value = ""
            if row_idx <= data_last_row:
                raw_value = worksheet.cell(row=row_idx, column=1).value
                value = str(raw_value).strip() if raw_value is not None else ""
                if _is_total_text(value):
                    value = ""
            if value and value == current_value:
                continue
            if merge_start is not None and row_idx - merge_start > 1:
                worksheet.merge_cells(
                    start_row=merge_start,
                    start_column=1,
                    end_row=row_idx - 1,
                    end_column=1,
                )
                worksheet.cell(row=merge_start, column=1).alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                )
            merge_start = row_idx if value else None
            current_value = value


def write_output_workbook(
    output_sheets: dict[str, pd.DataFrame],
    output_path: Path,
    outputs_dir: Path,
    report_title: str,
    header_cfg: dict,
    styling_cfg: dict,
    source_df: pd.DataFrame,
    period_text_override: str | None = None,
    sheet_layouts: dict[str, str] | None = None,
    sheet_options: dict[str, dict] | None = None,
) -> None:
    try:
        output_relative = output_path.resolve().relative_to(outputs_dir.resolve())
    except ValueError as exc:
        raise ValueError(
            "Path output tidak valid: file wajib berada di folder outputs/."
        ) from exc

    try:
        safe_output_path = resolve_runtime_relative_path(
            outputs_dir.parent,
            f"outputs/{output_relative.as_posix()}",
            root_name="outputs",
        )
    except ValueError as exc:
        raise ValueError(f"Path output tidak valid: {exc}") from exc

    safe_output_path.parent.mkdir(parents=True, exist_ok=True)
    source_period_text = source_df.attrs.get("period_text")
    if not isinstance(source_period_text, str) or not source_period_text.strip():
        source_period_text = None
    period_text = period_text_override or source_period_text or _build_period_text(source_df, header_cfg)
    used_sheet_names: set[str] = set()

    with pd.ExcelWriter(safe_output_path, engine="openpyxl") as writer:
        for raw_sheet_name, frame in output_sheets.items():
            sheet_name = sanitize_sheet_name(raw_sheet_name, used_sheet_names)
            layout_mode = (sheet_layouts or {}).get(raw_sheet_name, "standard")
            is_summary_sheet = (
                layout_mode == "plain" and raw_sheet_name.casefold().startswith("data")
            )
            if layout_mode == "standard":
                startrow = 3
                freeze_pane_default = "A5"
            elif layout_mode == "plain":
                startrow = 3 if is_summary_sheet else 0
                freeze_pane_default = "B5" if is_summary_sheet else "A2"
            else:
                raise ValueError(f"Mode layout sheet tidak didukung: '{layout_mode}'.")

            # Ekstrak _row_type sebelum ditulis ke Excel
            row_types: list[str] | None = None
            if "_row_type" in frame.columns:
                row_types = frame["_row_type"].tolist()
                frame = frame.drop(columns=["_row_type"])

            frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
            worksheet = writer.sheets[sheet_name]

            if layout_mode == "standard":
                worksheet["A1"] = report_title
                worksheet["A2"] = period_text
                worksheet["A3"] = f"Dibuat: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                worksheet["A1"].font = Font(
                    name=str(styling_cfg.get("font", "Calibri")),
                    bold=True,
                    size=14,
                )
                worksheet["A2"].font = Font(name=str(styling_cfg.get("font", "Calibri")))
                worksheet["A3"].font = Font(name=str(styling_cfg.get("font", "Calibri")))

            if is_summary_sheet:
                _apply_summary_sheet_style(
                    worksheet=worksheet,
                    frame=frame,
                    startrow=startrow,
                    styling_cfg=styling_cfg,
                    row_types=row_types,
                    sheet_options=(sheet_options or {}).get(raw_sheet_name),
                )
            else:
                _apply_worksheet_style(
                    worksheet=worksheet,
                    frame=frame,
                    startrow=startrow,
                    styling_cfg=styling_cfg,
                    freeze_pane_default=freeze_pane_default,
                    row_types=row_types,
                )
